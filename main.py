import sys
import tempfile
import os
import psutil
import json
from openpyxl import Workbook
from copy import copy
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QFileDialog,
    QVBoxLayout, QWidget, QMessageBox, QTableWidget, QTableWidgetItem, QProgressDialog
)
from PyQt6.QtGui import QColor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side, Font, alignment
from openpyxl.utils import get_column_letter
import re
from PyQt6.QtWidgets import QLabel, QLineEdit
from PyQt6.QtCore import pyqtSignal, Qt, QThread, QPropertyAnimation, QEasingCurve, QPoint, QParallelAnimationGroup
from PyQt6.QtWidgets import QSizePolicy
from PyQt6.QtWidgets import QScrollArea
from PyQt6.QtWidgets import QComboBox
import os



class DropFrame(QLabel):
    file_dropped = pyqtSignal(list)  # Сигнал с путем к файлу
    text_updated = pyqtSignal(str)


    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.text_updated.connect(self.setText)

        self.setCursor(Qt.CursorShape.OpenHandCursor)

        self.setText("Перетащите Excel-файл сюда / папку с файлами")
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #4CAF50;
                border-radius: 12px;
                padding: 15px 20px;
                font-size: 14px;
                color: #2E7D32;
                background-color: #F1F8E9;
                font-weight: 500;
            }

            QLabel:hover {
                background-color: #E8F5E9;
                border: 2px dashed #2E7D32;
            }
        """)
        self.setFixedHeight(100)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.label_file_path = QLabel("Файл не загружен")
        self.label_file_path.setStyleSheet("color: #388e3c; font-style: italic;")



    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

            self.setCursor(Qt.CursorShape.ClosedHandCursor)

            self.setStyleSheet("""
                QLabel {
                    border: 3px solid #2E7D32;
                    border-radius: 10px;
                    padding: 20px;
                    font-size: 16px;
                    color: #2E7D32;
                    background-color: #E8F5E9;
                }
            """)
        else:
            event.ignore()


    def dragLeaveEvent(self, event):

        self.setCursor(Qt.CursorShape.ClosedHandCursor)

        self.setStyleSheet("""
            DropFrame {
                background-color: #f0f0f0; 
                border: 2px dashed #66bb6a;
                border-radius: 10px;
            }
        """)

    def dropEvent(self, event):
        try:
            urls = event.mimeData().urls()
            all_file_paths = []

            for url in urls:
                file_path = url.toLocalFile()
                if os.path.isfile(file_path) and file_path.lower().endswith(('.xlsx', '.xls')):
                    all_file_paths.append(file_path)
                elif os.path.isdir(file_path):
                    for root, dirs, files in os.walk(file_path):
                        for file in files:
                            if file.lower().endswith(('.xlsx', '.xls')):
                                full_path = os.path.join(root, file)
                                all_file_paths.append(full_path)

            # Восстанавливаем стиль
            self.setStyleSheet("""
                QLabel {
                    border: 3px dashed #4CAF50;
                    border-radius: 10px;
                    padding: 40px;
                    font-size: 16px;
                    color: #555;
                    background-color: #F9FFF9;
                }
            """)

            if all_file_paths:
                unique_files = list(set(all_file_paths))
                self.file_dropped.emit(unique_files)
                self.setText("Перетащите Excel-файл сюда / папку с файлами")  # ← ВОЗВРАЩАЕМ ИСХОДНЫЙ ТЕКСТ

            event.acceptProposedAction()

        except Exception as e:
            print(f"Ошибка в dropEvent: {e}")
            self.setText("Ошибка при обработке")
            event.acceptProposedAction()


class ExcelApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("УЗСМК")
        self.setGeometry(100, 100, 1000, 600)

        import threading
        self.regex_lock = threading.Lock()  # ← Блокировка для regex

        self.df = None

        self.last_loaded_path = None  # Путь к последнему загруженному файлу

        # self.button_load = QPushButton("Загрузить Excel", self)
        # self.button_load.clicked.connect(self.load_excel)

        self.all_dropped_files = []

        self.regex_patterns = {
            'list_pattern': re.compile(r'-\d+[хx]\d+', re.IGNORECASE),
            'list_extract': re.compile(r'-(\d+)[хx]\d+', re.IGNORECASE),
            'pipe_remove': re.compile(r'труба', re.IGNORECASE),
            'brackets_content': re.compile(r'\([^)]*\)'),
            'brackets_only': re.compile(r'[\(\)]'),
            'bolt_pattern': re.compile(r'(болт.*?)-', re.IGNORECASE),
            'nut_pattern': re.compile(r'(гайка.*?)-', re.IGNORECASE),
            'washer_pattern': re.compile(r'шайба с(\d+)\.\d+', re.IGNORECASE),
            'channel_pattern': re.compile(r'\[(.*?)\]'),
            'non_letters': re.compile(r'[^a-zA-Zа-яА-Я]')
        }



        layout = QVBoxLayout()

        self.pending_files = []  # Добавьте эту строку для хранения отложенных файлов

        self.history = {
            'b5': [], 'c5': [], 'd5': [], 'f5': [], 'h5': [],
            'i5': [], 'j5': [], 'k5': [], 'l5': [], 'm5': [],
            'n5': [], 'o5': [], 'p5': [], 'q5': [], 'r5': []
        }


        self.label_b5, self.input_b5 = self.make_combo_with_label("Тип товара:", 'b5')
        self.label_c5, self.input_c5 = self.make_combo_with_label("Наименование Товара", 'c5')
        self.label_d5, self.input_d5 = self.make_combo_with_label("Наименование для ВОМ", 'd5')
        self.label_f5, self.input_f5 = self.make_combo_with_label("Нормативный документ", 'f5')
        self.label_h5, self.input_h5 = self.make_combo_with_label("Количество марок в единице товара", 'h5')
        self.label_i5, self.input_i5 = self.make_combo_with_label("Длина", 'i5')
        self.label_j5, self.input_j5 = self.make_combo_with_label("Ширина", 'j5')
        self.label_k5, self.input_k5 = self.make_combo_with_label("Высота", 'k5')
        self.label_l5, self.input_l5 = self.make_combo_with_label("Вид ГМ", 'l5')
        self.label_m5, self.input_m5 = self.make_combo_with_label("Краска1", 'm5')
        self.label_n5, self.input_n5 = self.make_combo_with_label("Цвет1", 'n5')
        self.label_o5, self.input_o5 = self.make_combo_with_label("Расход краски1, кг", 'o5')
        self.label_p5, self.input_p5 = self.make_combo_with_label("Краска2", 'p5')
        self.label_q5, self.input_q5 = self.make_combo_with_label("Цвет2", 'q5')
        self.label_r5, self.input_r5 = self.make_combo_with_label("Расход краски2, кг", 'r5')

        # Добавляем виджеты в layout
        for label, combo in [
            (self.label_b5, self.input_b5),
            (self.label_c5, self.input_c5),
            (self.label_d5, self.input_d5),
            (self.label_f5, self.input_f5),
            (self.label_h5, self.input_h5),
            (self.label_i5, self.input_i5),
            (self.label_j5, self.input_j5),
            (self.label_k5, self.input_k5),
            (self.label_l5, self.input_l5),
            (self.label_m5, self.input_m5),
            (self.label_n5, self.input_n5),
            (self.label_o5, self.input_o5),
            (self.label_p5, self.input_p5),
            (self.label_q5, self.input_q5),
            (self.label_r5, self.input_r5),
        ]:
            layout.addWidget(label)
            layout.addWidget(combo)

        self.button_load = QPushButton("Загрузить файл(ы)/папку", self)

        self.button_load.clicked.connect(self.universal_load)


        # В __init__:
        self.button_process_all = QPushButton("Обработать все файлы", self)

        self.button_process_all.clicked.connect(self.process_all_files)




        # ДОБАВЬТЕ ЭТО:
        # self.button_select_folder = QPushButton("Выбрать папку с файлами", self)
        # self.button_select_folder.clicked.connect(self.select_folder)
        # self.button_select_folder.setStyleSheet("""
        #     QPushButton {
        #         background-color: #ff9800;
        #         color: white;
        #         border-radius: 5px;
        #         padding: 8px 12px;
        #     }
        #     QPushButton:hover {
        #         background-color: #f57c00;
        #     }
        # """)



        self.button_load.setToolTip("Можно выбрать: один файл, несколько файлов или папку с файлами")

        self.setLayout(layout)

        # В __init__ после создания кнопок:
        self.button_clear_queue = QPushButton("Очистить очередь", self)

        self.button_clear_queue.clicked.connect(self.clear_file_queue)




        # Метка для отображения загруженного файла
        self.label_file_path = QLabel("Файл не загружен")
        self.label_file_path.setStyleSheet("color: #388e3c; font-weight: bold;")
        self.label_file_path.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # DropFrame с ограниченной высотой
        drop_frame = DropFrame()
        drop_frame.setFixedHeight(100)
        drop_frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        drop_frame.file_dropped.connect(self.handle_dropped_files)

        self.label_queue_status = QLabel("Файлов в очереди: 0")
        self.label_queue_status.setStyleSheet("""
            QLabel {
                color: #2E7D32;
                font-weight: bold;
                font-size: 13px;
                padding: 10px;
                border-radius: 6px;
                border: 1px solid #C8E6C9;
            }
        """)
        self.label_queue_status.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Главный вертикальный layout
        main_layout = QVBoxLayout()
        main_layout.addWidget(drop_frame)

        main_layout.addWidget(self.label_queue_status)

        # Layout с полями и подписями
        form_layout = QVBoxLayout()
        form_layout.addWidget(self.label_b5)
        form_layout.addWidget(self.input_b5)

        form_layout.addWidget(self.label_c5)
        form_layout.addWidget(self.input_c5)

        form_layout.addWidget(self.label_d5)
        form_layout.addWidget(self.input_d5)

        form_layout.addWidget(self.label_f5)
        form_layout.addWidget(self.input_f5)

        form_layout.addWidget(self.label_h5)
        form_layout.addWidget(self.input_h5)

        form_layout.addWidget(self.label_i5)
        form_layout.addWidget(self.input_i5)

        form_layout.addWidget(self.label_j5)
        form_layout.addWidget(self.input_j5)

        form_layout.addWidget(self.label_k5)
        form_layout.addWidget(self.input_k5)

        form_layout.addWidget(self.label_l5)
        form_layout.addWidget(self.input_l5)

        form_layout.addWidget(self.label_m5)
        form_layout.addWidget(self.input_m5)

        form_layout.addWidget(self.label_n5)
        form_layout.addWidget(self.input_n5)

        form_layout.addWidget(self.label_o5)
        form_layout.addWidget(self.input_o5)

        form_layout.addWidget(self.label_p5)
        form_layout.addWidget(self.input_p5)

        form_layout.addWidget(self.label_q5)
        form_layout.addWidget(self.input_q5)

        form_layout.addWidget(self.label_r5)
        form_layout.addWidget(self.input_r5)

        form_layout.addWidget(self.button_load)


        form_layout.addWidget(self.button_process_all)
        form_layout.addWidget(self.button_clear_queue)



        form_layout.addWidget(self.label_file_path)

        # Добавляем form_layout внутрь main_layout
        main_layout.addLayout(form_layout)

        # Устанавливаем основной layout
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        # В __init__ замени setStyleSheet на это:
        self.setStyleSheet("""
            QWidget {
                background-color: #E8F5E9;
                font-family: 'Segoe UI', Arial;
                color: #212121;
            }

            QPushButton {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
        stop: 0 #4CAF50, stop: 1 #45a049);
    color: white;
    border: none;
    border-radius: 6px;
    padding: 10px 16px;
    font-size: 13px;
    font-weight: bold;
    min-width: 120px;
    transition: all 0.3s ease;
}

QPushButton:hover {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
        stop: 0 #45a049, stop: 1 #3d8b40);
    box-shadow: 0 0 15px rgba(76, 175, 80, 0.6);
    transform: translateY(-1px);
}

QPushButton:pressed {
    background: #2E7D32;
    box-shadow: 0 0 8px rgba(76, 175, 80, 0.4);
    transform: translateY(1px);
}

            /* ПОЛЯ ВВОДА */
            QComboBox, QLineEdit {
                background-color: white;
                border: 1px solid #C8E6C9;
                border-radius: 4px;
                padding: 8px;
                font-size: 13px;
                min-width: 200px;
                selection-background-color: #4CAF50;
            }

            QComboBox:editable, QLineEdit:focus {
                border: 1px solid #4CAF50;
            }

            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid #C8E6C9;
            }
            
            
            

            /* МЕТКИ */
            QLabel {
                font-weight: 600;
                color: #2E7D32;
                font-size: 13px;
                padding: 4px 0;
            }

            /* SCROLLAREA */
            QScrollArea {
                border: none;
                background-color: #E8F5E9;
            }

            QScrollBar:vertical {
                border: none;
                background: #C8E6C9;
                width: 12px;
                margin: 0px;
            }

            QScrollBar::handle:vertical {
                background: #4CAF50;
                border-radius: 6px;
                min-height: 30px;
            }

            QScrollBar::handle:vertical:hover {
                background: #388E3C;
            }

            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }

            QScrollBar:horizontal {
                height: 0px;
            }
            
            /* КОМБОБОКСЫ С ИСТОРИЕЙ */
QComboBox {
    background-color: white;
    border: 1px solid #C8E6C9;
    border-radius: 4px;
    padding: 8px;
    font-size: 13px;
    min-width: 250px;
    selection-background-color: #4CAF50;
}

QComboBox:editable {
    background: white;
}

QComboBox:!editable, QComboBox::drop-down:editable {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
        stop: 0 #F1F8E9, stop: 0.4 #F1F8E9, stop: 0.5 #E8F5E9, stop: 1.0 #E8F5E9);
}

QComboBox:!editable:on, QComboBox::drop-down:editable:on {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
        stop: 0 #E8F5E9, stop: 0.4 #E8F5E9, stop: 0.5 #D7E8D7, stop: 1.0 #D7E8D7);
}

/* ПРОСТОЙ СТИЛЬ С СИСТЕМНОЙ СТРЕЛКОЙ */
QComboBox::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 20px;
    border-left: 1px solid #4CAF50;
}

QComboBox::down-arrow {
    image: url(none);
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 6px solid #2E7D32;
    width: 0;
    height: 0;
}

QComboBox::down-arrow:on {
    border-top: 6px solid #1B5E20;
}

/* ВЫПАДАЮЩИЙ СПИСОК */
QComboBox QAbstractItemView {
    border: 1px solid #4CAF50;
    border-radius: 4px;
    background-color: white;
    selection-background-color: #4CAF50;
    selection-color: white;
    outline: 0;
}

QComboBox QAbstractItemView::item {
    padding: 6px;
    border-bottom: 1px solid #E8F5E9;
}

QComboBox QAbstractItemView::item:selected {
    background-color: #4CAF50;
    color: white;
}

/* ПОДСКАЗКИ */
QToolTip {
    background-color: #2E7D32;
    color: white;
    border: 1px solid #1B5E20;
    border-radius: 4px;
    padding: 4px 8px;
}
        """)


        scroll_content = QWidget()
        scroll_content.setLayout(main_layout)
        scroll_area.setWidget(scroll_content)

        # Устанавливаем scroll_area как центральный виджет
        self.setCentralWidget(scroll_area)

        self.processing_thread = None
        self.progress_dialog = None

        self.setup_history_file()

    def showEvent(self, event):
        """Простая анимация прозрачности"""
        super().showEvent(event)

        self.setWindowOpacity(0)
        self.animation = QPropertyAnimation(self, b"windowOpacity")
        self.animation.setDuration(500)
        self.animation.setStartValue(0)
        self.animation.setEndValue(1)
        self.animation.start()

    def process_value(self, text):
        text = str(text).strip()
        lower_text = text.lower()

        # КРУГ - самый высокий приоритет (первый символ Ø)
        if text.startswith('Ø') or text.startswith('ø') or text.startswith('⌀'):
            return text, "Круг"

        # БОЛТ и ГАЙКА с приоритетом
        if 'болт' in lower_text and '-' in text:
            return text, "Болт"

        if 'гайка' in lower_text and '-' in text:
            return text, "Гайка"

        if 'труба' in lower_text:
            return text, "Труба"

        if 'шайба' in lower_text:
            return text, "Шайба"

            # ЛИСТ - ТОЛЬКО строгий паттерн: -числохчисло
            # Пример: -6х2235, -10х1000 и т.д.
        if (re.search(r'-\d+[хx]\d+', text, re.IGNORECASE) and
            'болт' not in lower_text and
            'гайка' not in lower_text and
            'бобышка' not in lower_text and
            'пробка' not in lower_text and
            'шпилька' not in lower_text):
            return text, "Лист"

        if text.startswith(('L', 'l')):
            return text, "Уголок"

        if text.startswith('['):
            return text, "Швеллер"

            # Последний случай - первое слово без цифр
        words = text.split()
        if words:
            first_word = words[0]
            # Просто убираем все не-буквы из первого слова
            clean_word = ''.join(char for char in first_word if char.isalpha())
            return text, clean_word or first_word

        return text, "Unknown"



    def setup_history_file(self):
        """Инициализирует путь к файлу истории"""
        try:
            if getattr(sys, 'frozen', False):
                # Если это exe-файл
                base_path = os.path.dirname(sys.executable)
            else:
                # Если это скрипт Python
                base_path = os.path.dirname(os.path.abspath(__file__))

            self.history_file = os.path.join(base_path, "app_history.json")
            print(f"Файл истории: {self.history_file}")  # Для отладки
            self.load_history()  # Загружаем историю при запуске

        except Exception as e:
            print(f"Ошибка инициализации файла истории: {e}")
            # Создаем запасной путь чтобы избежать None
            self.history_file = "app_history.json"


    def make_combo_with_label(self, label_text, history_key):
        label = QLabel(label_text)
        combo = QComboBox()
        combo.setEditable(True)
        combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        combo.setMaxVisibleItems(10)
        combo.setMinimumWidth(250)  # ← ДОБАВЬ ЭТУ СТРОКУ
        combo.setToolTip("Нажмите для выбора из истории")  # ← ДОБАВЬ ПОДСКАЗКУ


        line_edit = combo.lineEdit()
        if line_edit:
            line_edit.editingFinished.connect(lambda key=history_key, cb=combo: self.save_to_history(cb, key))

        combo.addItems(self.history[history_key])
        return label, combo

    def get_auto_value(self, text):
        """Автоматическое заполнение для 3-й строки по тексту из 1-й строки"""
        text = str(text).strip()

        # Словарь правил (сначала специфичные, потом общие)
        rules = {
            # Фланцы
            r'Фланец переходной.*': '09Г2С ФП-2024-КМД',
            r'Фланец \d-\d+-\d+': '09Г2С ГОСТ 28759.2',
            r'Фланец \d+-\d+-\d+-\d+-\w': '09Г2С ГОСТ33259-2015',
            r'Крышка \d-\d+-\d+': '09Г2С ОСТ 26-2008-83',
            r'Заглушка \d-\d+-\d+': '09Г2С АТК24.200.02-90',

            # Прокладки
            r'Прокладка СНП.*': 'ГОСТ Р 52376-2005',
            r'Прокладка \d+-\w+': 'ГОСТ 28759.6',
            r'Прокладка \w-\d+-\d+-\w+': 'ГОСТ 15180',
            r'Прокладка \d+-\ПМБ': 'ОСТ26.260.460-99',

            # Детали
            r'Бобышка БПО.*': '09Г2С ОСТ26.260.460-99',
            r'Пробка .*': '09Г2С ОСТ26.260.460-99',
            r'Муфта \d+': '09Г2С ГОСТ 8966-75',
            r'Сгон \d+': '09Г2С ГОСТ8969-75',
            r'Ниппель \d+': '09Г2С ГОСТ8967-75',
            r'Отвод.*': '09Г2С ГОСТ 17375-2001',
            r'Переход.*': '09Г2С ГОСТ 17378-2001',
            r'Тройник.*': '09Г2С ГОСТ 17376-2001',

            # Крепеж
            r'Болт.*': 'ГОСТ 7798-70',
            r'Гайка.*': 'ГОСТ 5915-70',
            r'Шайба \d+': 'ГОСТ 11371-78',
            r'Шпилька.*': '09Г2С ГОСТ9066-75',
            r'Шплинт.*': 'ГОСТ 397-79',
            r'Электроды.*': 'ГОСТ9467-75',
            r'Скоба.*': 'С245 ГОСТ17314-81',
            r'Штырь.*': 'С245 ГОСТ17314-81'
        }

        # Ищем совпадение по правилам
        for pattern, auto_value in rules.items():
            if re.search(pattern, text, re.IGNORECASE):
                return auto_value

        return ""  # Если не нашли совпадение


    def load_history(self):
        """Загружает историю из файла при запуске"""
        try:
            if os.path.exists(self.history_file):
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    saved_history = json.load(f)
                    # Обновляем историю для каждого поля
                    for key in self.history.keys():
                        if key in saved_history:
                            self.history[key] = saved_history[key]
                            # Обновляем комбобоксы
                            combo = getattr(self, f'input_{key}')
                            combo.clear()
                            combo.addItems(self.history[key])
        except Exception as e:
            print(f"Ошибка загрузки истории: {e}")

    def save_to_history(self, combo: QComboBox, key: str):
        text = combo.currentText().strip()
        if not text:
            return

        # Временно отключаем автосохранение чтобы избежать рекурсии
        hist = self.history[key]
        if text in hist:
            hist.remove(text)
        hist.insert(0, text)
        if len(hist) > 10:
            hist.pop()

        # Блокируем сигналы на время обновления комбобокса
        combo.blockSignals(True)
        combo.clear()
        combo.addItems(hist)
        combo.setCurrentText(text)
        combo.blockSignals(False)

        # Сохраняем историю (теперь безопасно)
        self.save_history_safe()

    def save_history_safe(self):
        """Безопасное сохранение истории без рекурсии"""
        try:
            if not hasattr(self, 'history_file') or self.history_file is None:
                # Если по какой-то причине history_file не инициализирован
                self.setup_history_file()

            # Сохраняем текущую историю
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, ensure_ascii=False, indent=2)
            print(f"История сохранена в: {self.history_file}")
        except Exception as e:
            print(f"Ошибка сохранения истории: {e}")


    def update_all_histories(self):
        """Только обновляет историю из текущих значений, без сохранения"""
        fields_mapping = {
            'b5': self.input_b5,
            'c5': self.input_c5,
            'd5': self.input_d5,
            'f5': self.input_f5,
            'h5': self.input_h5,
            'i5': self.input_i5,
            'j5': self.input_j5,
            'k5': self.input_k5,
            'l5': self.input_l5,
            'm5': self.input_m5,
            'n5': self.input_n5,
            'o5': self.input_o5,
            'p5': self.input_p5,
            'q5': self.input_q5,
            'r5': self.input_r5
        }

        for key, combo in fields_mapping.items():
            text = combo.currentText().strip()
            if text:
                hist = self.history[key]
                if text in hist:
                    hist.remove(text)
                hist.insert(0, text)
                if len(hist) > 10:
                    hist.pop()

    def closeEvent(self, event):
        """Сохраняет историю при закрытии приложения"""
        self.save_history()
        event.accept()


    def load_excel_from_path(self, file_path):
        self.last_loaded_path = file_path  # сохраняем путь

        try:
            df_raw = pd.read_excel(file_path, header=None)

            new_header = [
                "№", "Тип Товара", "Наименование Товара", "Наименование для ВОМ",
                "Марка", "Нормативный документ", "Вес шт, кг", "Количество марок в единице Товара",
                "Длина", "Ширина", "Высота", "Вид ГМ", "Краска1", "Цвет1", "Расход краски1, кг",
                "Краска2", "Цвет2", "Расход краски2, кг"
            ]

            base_df = pd.DataFrame([new_header], columns=new_header)
            base_df = pd.concat([base_df, pd.DataFrame([[""] * len(new_header)] * 2, columns=new_header)],
                                ignore_index=True)

            l_vals = df_raw.iloc[2:, 11].dropna().tolist() if df_raw.shape[1] > 11 else []
            e_vals = df_raw.iloc[2:, 4].dropna().tolist() if df_raw.shape[1] > 4 else []
            f_vals = df_raw.iloc[2:, 5].dropna().tolist() if df_raw.shape[1] > 5 else []

            max_len = max(len(l_vals), len(e_vals), len(f_vals))

            while len(base_df.columns) < 19 + max_len:
                base_df[f"Доп.{len(base_df.columns) - 18}"] = ""

            self.df = base_df.copy()

            while len(self.df) < 4:
                self.df.loc[len(self.df)] = [""] * len(self.df.columns)

            for idx, val in enumerate(f_vals):
                self.df.iat[0, 18 + idx] = val  # строка 1

            for idx, val in enumerate(l_vals):
                self.df.iat[2, 18 + idx] = val  # строка 3

            for idx, val in enumerate(e_vals):
                self.df.iat[3, 18 + idx] = val  # строка 4

            self.label_file_path.setText(f"Загружен файл: {os.path.basename(file_path)}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))



    def load_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите Excel-файл", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.last_loaded_path = file_path
            self.label_file_path.setText(f"Загружен файл: {os.path.basename(file_path)}")
            self.load_excel_from_path(file_path)

    def handle_dropped_files(self, file_paths):
        """Обработчик перетащенных файлов с накопительным эффектом"""
        try:
            # Добавляем новые файлы к существующим
            for path in file_paths:
                if path not in self.all_dropped_files:
                    self.all_dropped_files.append(path)

            # Обновляем статус
            total_files = len(self.all_dropped_files)
            self.update_file_status(f"Файлов в очереди: {total_files}")
            self.label_queue_status.setText(f"Файлов в очереди: {total_files}")

            # Обновляем текст в DropFrame
            drop_frame = self.findChild(DropFrame)
            if drop_frame:
                drop_frame.setText(f"Файлов в очереди: {total_files}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки: {str(e)}")

    def get_output_path(self):
        """Получаем путь для сохранения"""
        default_path = os.path.join(os.path.expanduser("~"), "Desktop", "обработанные_файлы.xlsx")

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить результат", default_path, "Excel Files (*.xlsx)"
        )

        return file_path if file_path else None



    def update_file_status(self, text):
        """Обновляет статус загруженного файла"""
        if hasattr(self, 'label_file_path'):
            self.label_file_path.setText(text)
        else:
            print(f"Статус: {text}")



    def check_memory_issues(self):
        """Проверка на утечки памяти"""
        process = psutil.Process()
        print(f"Используется памяти: {process.memory_info().rss / 1024 / 1024:.2f} MB")

    def on_processing_finished(self, success, message):
        """Обработка завершения"""
        self.progress_dialog.close()

        if success:
            QMessageBox.information(self, "Готово", f"Файл сохранен: {message}")
        else:
            QMessageBox.critical(self, "Ошибка", message)

    def on_processing_error(self, error_message):
        """Обработка отдельных ошибок файлов"""
        print(error_message)  # или можно собирать в лог

    def apply_user_data(self, user_data):
        """Применяет пользовательские данные к интерфейсу"""
        if not user_data:
            return

        # Устанавливаем значения для всех полей ввода
        if 'b5' in user_data:
            self.input_b5.setCurrentText(user_data['b5'])

        if 'c5' in user_data:
            self.input_c5.setCurrentText(user_data['c5'])

        if 'd5' in user_data:
            self.input_d5.setCurrentText(user_data['d5'])

        if 'f5' in user_data:
            self.input_f5.setCurrentText(user_data['f5'])

        if 'h5' in user_data:
            self.input_h5.setCurrentText(user_data['h5'])

        if 'i5' in user_data:
            self.input_i5.setCurrentText(user_data['i5'])

        if 'j5' in user_data:
            self.input_j5.setCurrentText(user_data['j5'])

        if 'k5' in user_data:
            self.input_k5.setCurrentText(user_data['k5'])

        if 'l5' in user_data:
            self.input_l5.setCurrentText(user_data['l5'])

        if 'm5' in user_data:
            self.input_m5.setCurrentText(user_data['m5'])

        if 'n5' in user_data:
            self.input_n5.setCurrentText(user_data['n5'])

        if 'o5' in user_data:
            self.input_o5.setCurrentText(user_data['o5'])

        if 'p5' in user_data:
            self.input_p5.setCurrentText(user_data['p5'])

        if 'q5' in user_data:
            self.input_q5.setCurrentText(user_data['q5'])

        if 'r5' in user_data:
            self.input_r5.setCurrentText(user_data['r5'])

    def get_user_data_for_batch(self):
        """Получает текущие данные пользователя для пакетной обработки"""
        return {
            'b5': self.input_b5.currentText().strip(),
            'c5': self.input_c5.currentText().strip(),
            'd5': self.input_d5.currentText().strip(),
            'f5': self.input_f5.currentText().strip(),
            'h5': self.input_h5.currentText().strip(),
            'i5': self.input_i5.currentText().strip(),
            'j5': self.input_j5.currentText().strip(),
            'k5': self.input_k5.currentText().strip(),
            'l5': self.input_l5.currentText().strip(),
            'm5': self.input_m5.currentText().strip(),
            'n5': self.input_n5.currentText().strip(),
            'o5': self.input_o5.currentText().strip(),
            'p5': self.input_p5.currentText().strip(),
            'q5': self.input_q5.currentText().strip(),
            'r5': self.input_r5.currentText().strip()
        }

    def confirm_batch_processing(self, file_count):
        """Запрос подтверждения для пакетной обработки"""
        user_data = self.get_user_data_for_batch()

        # Формируем текст с данными
        data_text = "\n".join([f"{key}: {value}" for key, value in user_data.items() if value])

        reply = QMessageBox.question(
            self,
            "Подтверждение пакетной обработки",
            f"Обработать {file_count} файлов со следующими данными?\n\n"
            f"{data_text}\n\n"
            "Продолжить?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )

        return user_data if reply == QMessageBox.StandardButton.Yes else None

    def load_excel_file(self, file_path, silent=False):
        """Загрузка файла с преобразованием данных"""
        try:
            # Вместо прямой загрузки - используем ваш метод преобразования
            self.load_excel_from_path(file_path)  # ← ЭТО КЛЮЧЕВОЕ ИЗМЕНЕНИЕ

            if not silent:
                self.update_file_status(f"Загружен: {os.path.basename(file_path)}")
                QMessageBox.information(self, "Успех", f"Файл загружен: {os.path.basename(file_path)}")

            return True

        except Exception as e:
            if not silent:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
            return False

    def select_folder(self):
        """Выбор папки с файлами"""
        try:
            folder_path = QFileDialog.getExistingDirectory(
                self,
                "Выберите папку с Excel-файлами",
                os.path.expanduser("~")
            )

            if not folder_path:
                return

            # Находим все Excel-файлы в папке
            excel_files = []
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith(('.xlsx', '.xls')):
                        excel_files.append(os.path.join(root, file))

            if not excel_files:
                QMessageBox.information(self, "Информация", "В папке не найдено Excel-файлов")
                return

            if len(excel_files) == 1:
                # Если один файл - загружаем его
                self.load_excel_file(excel_files[0])
            else:
                # Если несколько файлов - запускаем обработку
                self.process_multiple_files(excel_files)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при выборе папки: {str(e)}")







    def start_processing(self, file_paths):
        """Запуск обработки файлов в потоке"""
        if not file_paths:
            return

        user_data = self.get_user_data_for_batch()
        if not user_data:
            return

        self.processing_thread = FileProcessingThread(
            file_paths,
            user_data,
            self.regex_patterns,  # ← ПЕРЕДАЕМ REGEX
            self.regex_lock  # ← ПЕРЕДАЕМ БЛОКИРОВКУ
        )

        # Создаем прогресс-диалог
        self.progress_dialog = QProgressDialog(
            "Обработка файлов...",
            "Отмена",
            0,
            len(file_paths),
            self
        )
        self.progress_dialog.setWindowTitle("Обработка")
        self.progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        self.progress_dialog.canceled.connect(self.cancel_processing)

        # Создаем и запускаем поток
        self.processing_thread = FileProcessingThread(file_paths, user_data)
        self.processing_thread.progress.connect(self.update_progress)
        self.processing_thread.finished.connect(self.on_processing_finished)
        self.processing_thread.error.connect(self.on_processing_error)

        self.processing_thread.start()
        self.progress_dialog.show()

    def process_multiple_files(self, file_paths):
        """Обработка нескольких файлов с объединением в один лист"""
        try:
            if not file_paths:
                return

            user_data = self.get_user_data_for_batch()
            output_path = self.get_output_path()
            if not output_path:
                return

            # Шаг 1: Сбор всех уникальных заголовков из столбцов S+
            all_s_headers = set()
            all_tables_data = []  # Будем сохранять данные всех таблиц

            for file_path in file_paths:
                try:
                    # Обрабатываем ВСЕ листы файла
                    xl_file = pd.ExcelFile(file_path)
                    file_name = os.path.splitext(os.path.basename(file_path))[0]

                    for sheet_name in xl_file.sheet_names:
                        try:
                            # Обрабатываем конкретный лист
                            success, table_data = self.process_single_sheet_for_combined(
                                file_path, sheet_name, user_data
                            )

                            if success:
                                all_tables_data.append({
                                    'file_name': file_name,
                                    'sheet_name': sheet_name,
                                    'data': table_data
                                })

                                # Собираем заголовки из столбцов S+ (индексы 18+)
                                if len(table_data) > 0:
                                    for col_idx in range(18, len(table_data[0])):
                                        header = table_data[0][col_idx]
                                        if header and str(header).strip():
                                            all_s_headers.add(header)

                                print(f"✓ Обработан лист: {sheet_name} в файле {file_name}")
                            else:
                                print(f"✗ Ошибка листа: {sheet_name} в файле {file_name}")

                        except Exception as e:
                            print(f"❌ Ошибка обработки листа {sheet_name}: {e}")
                            continue

                    print(f"✓ Обработан файл: {os.path.basename(file_path)}")

                except Exception as e:
                    print(f"❌ Ошибка открытия файла {file_path}: {e}")
                    continue

            # Выводим собранные заголовки для проверки
            print("Собранные заголовки столбцов S+:")
            for header in sorted(all_s_headers):
                print(f"  - {header}")

            # Если нет данных, выходим
            if not all_tables_data:
                QMessageBox.warning(self, "Нет данных", "Не удалось обработать ни одной таблицы.")
                return

            # Создаем новую книгу с одним листом
            wb = Workbook()
            ws = wb.active
            ws.title = "Объединенные_данные"

            current_row = 1  # Текущая строка для вставки

            # Шаг 2: Заполняем столбцы A-R
            for i, table_info in enumerate(all_tables_data):
                table_data = table_info['data']
                if i == 0:
                    # Первая таблица - вставляем полностью (все строки)
                    for row_idx, row_data in enumerate(table_data):
                        # Копируем только столбцы A-R (индексы 0-17)
                        for col_idx in range(0, 18):
                            if col_idx < len(row_data) and row_data[col_idx] is not None:
                                ws.cell(row=current_row + row_idx, column=col_idx + 1, value=row_data[col_idx])
                    current_row += len(table_data)
                else:
                    # Последующие таблицы - только 5-я строка (индекс 4) для A-R
                    if len(table_data) > 4:
                        row_data = table_data[4]
                        for col_idx in range(0, 18):
                            if col_idx < len(row_data) and row_data[col_idx] is not None:
                                ws.cell(row=current_row, column=col_idx + 1, value=row_data[col_idx])
                        current_row += 1

            first_table_data = all_tables_data[0]['data']
            s_columns_data = {}  # Будет хранить данные для каждого столбца S+

            # Проходим по всем столбцам S+ первой таблицы
            for col_idx in range(18, len(first_table_data[0])):
                header = first_table_data[0][col_idx]  # Заголовок из первой строки
                if header and str(header).strip():
                    # Сохраняем все 4 строки для этого столбца
                    column_data = []
                    for row_idx in range(4):  # Первые 4 строки
                        if col_idx < len(first_table_data[row_idx]):
                            column_data.append(first_table_data[row_idx][col_idx])
                        else:
                            column_data.append(None)
                    s_columns_data[header] = column_data

            # Создаем映射 заголовков к столбцам и заполняем все 4 строки
            sorted_headers = sorted(list(s_columns_data.keys()))
            header_to_col = {}

            for idx, header in enumerate(sorted_headers, start=19):  # Начинаем с столбца S
                header_to_col[header] = idx

                # Записываем все 4 строки шапки
                for row_idx, value in enumerate(s_columns_data[header]):
                    if value is not None:
                        ws.cell(row=row_idx + 1, column=idx, value=value)

            # Шаг 4: Заполняем данные для столбцов S+
            # Для первой таблицы заполняем 5-ю строку
            for col_idx in range(18, len(first_table_data[4])):
                header = first_table_data[0][col_idx]  # Заголовок из первой строки
                value = first_table_data[4][col_idx]  # Значение из 5-й строки
                if header and value is not None:
                    target_col = header_to_col.get(header)
                    if target_col:
                        ws.cell(row=5, column=target_col, value=value)

            # Для остальных таблиц заполняем только 5-ю строку
            current_data_row = 6  # Начинаем с 6-й строки (после первой таблицы)
            for table_info in all_tables_data[1:]:
                table_data = table_info['data']
                if len(table_data) > 4:
                    for col_idx in range(18, len(table_data[4])):
                        header = table_data[0][col_idx]  # Заголовок из первой строки
                        value = table_data[4][col_idx]  # Значение из 5-й строки
                        if header and value is not None:
                            target_col = header_to_col.get(header)
                            if target_col:
                                ws.cell(row=current_data_row, column=target_col, value=value)
                    current_data_row += 1

            # Добавляем этот код перед сохранением файла
            thick_border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )

            # Применяем жирные границы ко всем ячейкам
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thick_border

            # Добавляем столбец "Конец" в самый правый столбец
            max_col = ws.max_column
            end_col = max_col + 1

            # Записываем "Конец" в первую строку нового столбца
            ws.cell(row=1, column=end_col, value="Конец")
            end_cell = ws.cell(row=1, column=end_col)
            end_cell.font = Font(bold=True)  # Жирный шрифт
            end_cell.border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )

            # Сохраняем файл
            wb.save(output_path)

            QMessageBox.information(
                self, "Готово",
                f"Объединено {len(all_tables_data)} таблиц в один лист\n"
                f"Сохранено в: {output_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки: {str(e)}")

    def process_single_sheet_for_combined(self, file_path, sheet_name, user_data):
        """Обрабатывает один лист и возвращает данные в виде списка строк"""
        try:
            # 1. Загружаем конкретный лист
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

            # 2. Создаем временный DataFrame
            new_header = [
                "№", "Тип Товара", "Наименование Товара", "Наименование для ВОМ",
                "Марка", "Нормативный документ", "Вес шт, кг", "Количество марок в единице Товара",
                "Длина", "Ширина", "Высота", "Вид ГМ", "Краска1", "Цвет1", "Расход краски1, кг",
                "Краска2", "Цвет2", "Расход краски2, кг"
            ]

            base_df = pd.DataFrame([new_header], columns=new_header)
            base_df = pd.concat([base_df, pd.DataFrame([[""] * len(new_header)] * 2, columns=new_header)],
                                ignore_index=True)

            l_vals = df_raw.iloc[2:, 11].dropna().tolist() if df_raw.shape[1] > 11 else []
            e_vals = df_raw.iloc[2:, 4].dropna().tolist() if df_raw.shape[1] > 4 else []
            f_vals = df_raw.iloc[2:, 5].dropna().tolist() if df_raw.shape[1] > 5 else []

            max_len = max(len(l_vals), len(e_vals), len(f_vals))

            while len(base_df.columns) < 19 + max_len:
                base_df[f"Доп.{len(base_df.columns) - 18}"] = ""

            current_df = base_df.copy()

            while len(current_df) < 4:
                current_df.loc[len(current_df)] = [""] * len(current_df.columns)

            for idx, val in enumerate(f_vals):
                current_df.iat[0, 18 + idx] = val  # строка 1

            for idx, val in enumerate(l_vals):
                current_df.iat[2, 18 + idx] = val  # строка 3

            for idx, val in enumerate(e_vals):
                current_df.iat[3, 18 + idx] = val  # строка 4

            # 3. Временно сохраняем преобразованный DataFrame
            temp_dir = tempfile.gettempdir()
            temp_output = os.path.join(temp_dir, f"temp_process_{os.getpid()}_{sheet_name}.xlsx")
            current_df.to_excel(temp_output, index=False, header=False)

            # 4. Открываем оригинальный лист
            original_wb = load_workbook(file_path, data_only=True)
            original_ws = original_wb[sheet_name] if sheet_name in original_wb.sheetnames else original_wb.active

            # 5. Открываем временный файл
            wb = load_workbook(temp_output)
            ws = wb.active

            # 6. Применяем логику оформления (без дублирования!)
            # Цвета
            fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            fill_pink = PatternFill(start_color="D02090", end_color="D02090", fill_type="solid")
            fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            fill_darkblue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
            fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
            fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Чёрные границы
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            col_values = []
            for row in range(3, 23):  # от A3 до A22
                raw_value = original_ws.cell(row=row, column=1).value
                if raw_value is None:
                    continue

                # Преобразуем в строку и чистим от мусора
                cleaned = str(raw_value).replace('\n', ' ').replace('\r', '').replace('"', '').strip()

                if cleaned:  # добавляем только если реально есть содержимое
                    col_values.append(cleaned)

            # Склеиваем результат
            joined_string = "".join(col_values)
            ws["E5"].value = joined_string

            d5_value = user_data.get('d5', '').strip()

            # Объединяем D5 (из интерфейса) и E5 (joined_string) в D5
            if d5_value and joined_string:
                ws["D5"].value = f"{d5_value} {joined_string}"
            elif d5_value:
                ws["D5"].value = d5_value
            elif joined_string:
                ws["D5"].value = joined_string

            # Устанавливаем значения из user_data
            cell_mapping = {
                'b5': 'B5', 'c5': 'C5', 'f5': 'F5', 'h5': 'H5',
                'i5': 'I5', 'j5': 'J5', 'k5': 'K5', 'l5': 'L5',
                'm5': 'M5', 'n5': 'N5', 'o5': 'O5', 'p5': 'P5',
                'q5': 'Q5', 'r5': 'R5'
            }

            for data_key, cell_ref in cell_mapping.items():
                value = user_data.get(data_key, '').strip()
                if value:
                    ws[cell_ref].value = value

            sum_j = 0
            for row in range(3, 53):  # J3:J52
                raw_val = original_ws.cell(row=row, column=10).value  # колонка J = 10
                if raw_val is None:
                    continue
                try:
                    # Преобразуем в строку, убираем пробелы и "мм"
                    cleaned = str(raw_val).replace(" ", "").replace("мм", "").replace(",", ".").strip()
                    cleaned = re.sub(r"[^\d.]", "", cleaned)
                    if cleaned and cleaned != '.':  # ← Проверка на пустую строку и точку
                        number = float(cleaned)
                        sum_j += number
                except Exception as err:
                    print(f"Ошибка в строке {row}: {err}")
                    continue

            # Записываем сумму в G5
            ws["G5"].value = sum_j

            # Закрашиваем G5 болотным цветом
            fill_olive = PatternFill(start_color="808000", end_color="808000", fill_type="solid")
            ws["G5"].fill = fill_olive

            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(9, 12):
                ws.cell(row=4, column=col).value = "мм"

            ws["A5"].value = 1

            # O5
            try:
                g5_value = float(ws["G5"].value or 0)
                o5_value = float(user_data.get('o5', '0').replace(",", ".") or 0)
                ws["O5"].value = round(o5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете O5:", e)

            # R5
            try:
                g5_value = float(ws["G5"].value or 0)
                r5_value = float(user_data.get('r5', '0').replace(",", ".") or 0)
                ws["R5"].value = round(r5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете R5:", e)

            # Остальная логика обработки...
            # Обработка данных из столбцов K и L (перенос в столбцы S+)
            start_row_source = 3  # начинаем с 3-й строки исходного листа
            output_row = 3  # 3-я строка в целевом листе
            column_offset = 0  # сдвиг по столбцам, начиная со столбца S (19)
            empty_count = 0  # счётчик подряд пустых строк

            while empty_count < 3:
                k_value = original_ws.cell(row=start_row_source, column=11).value  # K = 11
                l_value = original_ws.cell(row=start_row_source, column=12).value  # L = 12

                if (k_value is None or str(l_value).strip() == "") and (l_value is None or str(k_value).strip() == ""):
                    empty_count += 1
                else:
                    l_str = str(l_value).strip() if l_value is not None else ""
                    k_str = str(k_value).strip() if k_value is not None else ""
                    combined = (k_str + " " + l_str).strip()
                    col_letter = get_column_letter(19 + column_offset)  # S = 19
                    ws[f"{col_letter}{output_row}"] = combined

                    column_offset += 1
                    empty_count = 0  # сброс, т.к. строка не пустая

                start_row_source += 1

            # Обработка данных из столбцов F, C, I (для строки 5)
            empty_count = 0
            i = 3  # начинаем с F3
            output_row = 5
            column_offset = 0

            while empty_count < 3:
                f_cell = original_ws.cell(row=i, column=5)  # F = 6
                f_value = f_cell.value

                if f_value is None or str(f_value).strip() == "":
                    empty_count += 1
                    i += 1
                    continue

                f_value_clean = str(f_value).strip().lower()

                if f_value_clean == "кг":
                    value_to_write = original_ws.cell(row=i, column=9).value  # I = 9
                    empty_count = 0
                elif f_value_clean == "шт":
                    value_to_write = original_ws.cell(row=i, column=3).value  # C = 3
                    empty_count = 0
                else:
                    # если что-то другое (не "кг" и не "шт"), считаем как пустую
                    empty_count += 1
                    i += 1
                    continue

                col_letter = get_column_letter(19 + column_offset)  # S = 19
                ws[f"{col_letter}{output_row}"] = value_to_write

                column_offset += 1
                i += 1

            # Автоматическое заполнение для столбцов с "ШТ" в 4-й строке
            start_row_source = 3  # начинаем с 3-й строки исходного листа
            empty_count = 0
            column_offset = 0

            for col_idx in range(19, max_col + 1):
                if col_idx <= ws.max_column:
                    cell_4th_row = ws.cell(row=4, column=col_idx).value
                    cell_1st_row = ws.cell(row=1, column=col_idx).value

                    # ЕСЛИ В 4-Й СТРОКЕ "ШТ" - АВТОМАТИЧЕСКОЕ ЗАПОЛНЕНИЕ
                    if cell_4th_row and str(cell_4th_row).strip().lower() == "шт" and cell_1st_row:
                        text_to_analyze = str(cell_1st_row).strip()
                        auto_value = ""

                        # ТАБЛИЦА СООТВЕТСТВИЙ
                        if re.search(r'Фланец переходной.*\d+-\d+.*\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С ФП-2024-КМД"
                        elif re.search(r'Фланец \d-\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 28759.2"
                        elif re.search(r'Фланец \d+-\d+-\d+-\d+-\w', text_to_analyze):
                            auto_value = "09Г2С ГОСТ33259-2015"
                        elif re.search(r'Крышка \d-\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С ОСТ 26-2008-83"
                        elif re.search(r'Заглушка \d-\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С АТК24.200.02-90"
                        elif re.search(r'Прокладка СНП', text_to_analyze):
                            auto_value = "ГОСТ Р 52376-2005"
                        elif re.search(r'Прокладка \d+-\w+', text_to_analyze):
                            auto_value = "ГОСТ 28759.6"
                        elif re.search(r'Прокладка \w-\d+-\d+-\w+', text_to_analyze):
                            auto_value = "ГОСТ 15180"
                        elif re.search(r'Прокладка +ПМБ-\d+', text_to_analyze):
                            auto_value = "ОСТ26.260.460-99"
                        elif re.search(r'(Бобышка БПО|Пробка)', text_to_analyze):
                            auto_value = "09Г2С ОСТ26.260.460-99"
                        elif re.search(r'Муфта \d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 8966-75"
                        elif re.search(r'Сгон \d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ8969-75"
                        elif re.search(r'Ниппель \d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ8967-75"
                        elif re.search(r'Отвод', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 17375-2001"
                        elif re.search(r'Переход', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 17378-2001"
                        elif re.search(r'Тройник', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 17376-2001"
                        elif re.search(r'Болт', text_to_analyze):
                            auto_value = "ГОСТ 7798-70"
                        elif re.search(r'Гайка', text_to_analyze):
                            auto_value = "ГОСТ 5915-70"
                        elif re.search(r'Шайба \d+', text_to_analyze):
                            auto_value = "ГОСТ 11371-78"
                        elif re.search(r'Шпилька', text_to_analyze):
                            auto_value = "09Г2С ГОСТ9066-75"
                        elif re.search(r'Шплинт', text_to_analyze):
                            auto_value = "ГОСТ 397-79"
                        elif re.search(r'Электроды', text_to_analyze):
                            auto_value = "ГОСТ9467-75"
                        elif re.search(r'(Скоба|Штырь)', text_to_analyze):
                            auto_value = "С245 ГОСТ17314-81"

                        ws.cell(row=3, column=col_idx).value = auto_value

                    # ИНАЧЕ - СТАРАЯ ЛОГИКА ПЕРЕНОСА ИЗ K И L
                    else:
                        k_value = original_ws.cell(row=start_row_source, column=11).value  # K = 11
                        l_value = original_ws.cell(row=start_row_source, column=12).value  # L = 12

                        if (k_value is None or str(k_value).strip() == "") and (
                                l_value is None or str(l_value).strip() == ""):
                            empty_count += 1
                        else:
                            k_str = str(k_value).strip() if k_value is not None else ""
                            l_str = str(l_value).strip() if l_value is not None else ""
                            combined = (k_str + " " + l_str).strip()

                            ws.cell(row=3, column=col_idx).value = combined

                            column_offset += 1
                            empty_count = 0

                        start_row_source += 1

            # Применение стилей оформления
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border

                    if row_idx == 1 and col_idx <= 19:
                        cell.fill = fill_blue  # A1:S1

                    if row_idx in [2, 3, 4] and col_idx in [9, 10, 11]:
                        cell.fill = fill_blue  # J2:L3

                    if row_idx == 1 and col_idx >= 19:
                        cell.fill = fill_pink  # S1 и далее

                    if row_idx == 3 and col_idx >= 19:
                        cell.fill = fill_orange  # S3 и далее

                    if row_idx == 4 and col_idx >= 19:
                        cell.fill = fill_darkblue  # S4 и далее

                    if row_idx == 1 and col_idx >= 19:
                        top_cell_value = ws.cell(row=1, column=col_idx).value
                        if top_cell_value:
                            value_row1, value_row2 = self.process_value(top_cell_value)
                            # Записываем в строку 1 и строку 2
                            ws.cell(row=1, column=col_idx).value = value_row1
                            ws.cell(row=2, column=col_idx).value = value_row2

            # 7. Собираем все данные из обработанного листа
            all_data = []
            max_row = ws.max_row
            max_col = ws.max_column

            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    row_data.append(cell.value)
                all_data.append(row_data)

            # 8. Очистка
            wb.close()
            original_wb.close()
            os.remove(temp_output)

            return True, all_data

        except Exception as e:
            print(f"Ошибка обработки листа {sheet_name}: {e}")
            import traceback
            traceback.print_exc()
            return False, []

    def make_columns_unique(self, worksheet):
        """Делает значения уникальными в столбцах, начиная с S"""
        # Определяем диапазон столбцов (S и далее)
        start_col = 19  # S - это 19-й столбец
        max_col = worksheet.max_column

        for col in range(start_col, max_col + 1):
            # Собираем все значения в столбце
            values = []
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    values.append(cell_value)

            # Оставляем только уникальные значения
            unique_values = list(dict.fromkeys(values))  # Сохраняем порядок

            # Записываем уникальные значения обратно в столбец
            for row, value in enumerate(unique_values, start=1):
                if row <= worksheet.max_row:
                    worksheet.cell(row=row, column=col).value = value
                else:
                    # Если уникальных значений больше, чем строк, добавляем новые строки
                    worksheet.cell(row=row, column=col).value = value

            # Очищаем оставшиеся ячейки, если уникальных значений меньше, чем было строк
            for row in range(len(unique_values) + 1, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col).value = None

    # def process_single_sheet(self, file_path, sheet_name, output_ws, user_data):
    #     """Обрабатывает один лист из файла"""
    #     try:
    #         # 1. Загружаем конкретный лист
    #         df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    #
    #         # ТВОЯ ЛОГИКА ПРЕОБРАЗОВАНИЯ DataFrame:
    #         new_header = [
    #             "№", "Тип Товара", "Наименование Товара", "Наименование для ВОМ",
    #             "Марка", "Нормативный документ", "Вес шт, кг", "Количество марок в единице Товара",
    #             "Длина", "Ширина", "Высота", "Вид ГМ", "Краска1", "Цвет1", "Расход краски1, кг",
    #             "Краска2", "Цвет2", "Расход краски2, кг"
    #         ]
    #
    #         base_df = pd.DataFrame([new_header], columns=new_header)
    #         base_df = pd.concat([base_df, pd.DataFrame([[""] * len(new_header)] * 2, columns=new_header)],
    #                             ignore_index=True)
    #
    #         l_vals = df_raw.iloc[2:, 11].dropna().tolist() if df_raw.shape[1] > 11 else []
    #         e_vals = df_raw.iloc[2:, 4].dropna().tolist() if df_raw.shape[1] > 4 else []
    #         f_vals = df_raw.iloc[2:, 5].dropna().tolist() if df_raw.shape[1] > 5 else []
    #
    #         max_len = max(len(l_vals), len(e_vals), len(f_vals))
    #
    #         while len(base_df.columns) < 19 + max_len:
    #             base_df[f"Доп.{len(base_df.columns) - 18}"] = ""
    #
    #         current_df = base_df.copy()
    #
    #         while len(current_df) < 4:
    #             current_df.loc[len(current_df)] = [""] * len(current_df.columns)
    #
    #         for idx, val in enumerate(f_vals):
    #             current_df.iat[0, 18 + idx] = val  # строка 1
    #
    #         for idx, val in enumerate(l_vals):
    #             current_df.iat[2, 18 + idx] = val  # строка 3
    #
    #         for idx, val in enumerate(e_vals):
    #             current_df.iat[3, 18 + idx] = val  # строка 4
    #
    #         # 2. Временно сохраняем преобразованный DataFrame
    #         temp_dir = tempfile.gettempdir()
    #         temp_output = os.path.join(temp_dir, f"temp_process_{os.getpid()}_{sheet_name}.xlsx")
    #         current_df.to_excel(temp_output, index=False, header=False)
    #
    #         # 3. Открываем оригинальный лист
    #         original_wb = load_workbook(file_path, data_only=True)
    #         original_ws = original_wb[sheet_name] if sheet_name in original_wb.sheetnames else original_wb.active
    #
    #         # 4. Открываем временный файл для оформления
    #         wb = load_workbook(temp_output)
    #         ws = wb.active
    #
    #         # 5. ПРИМЕНЯЕМ ЛОГИКУ ОФОРМЛЕНИЯ
    #         # Цвета
    #         fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # светло-голубой
    #         fill_pink = PatternFill(start_color="D02090", end_color="D02090", fill_type="solid")  # малиновый
    #         fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # оранжевый
    #         fill_darkblue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")  # тёмно-голубой
    #         fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  # коричневый
    #         fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # фиолетовый
    #         fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # желтый
    #
    #         # Чёрные границы
    #         border = Border(
    #             left=Side(style='thin', color='000000'),
    #             right=Side(style='thin', color='000000'),
    #             top=Side(style='thin', color='000000'),
    #             bottom=Side(style='thin', color='000000')
    #         )
    #
    #
    #
    #         col_values = []
    #         for row in range(3, 23):  # от A3 до A22
    #             raw_value = original_ws.cell(row=row, column=1).value
    #             if raw_value is None:
    #                 continue
    #
    #             # Преобразуем в строку и чистим от мусора
    #             cleaned = str(raw_value).replace('\n', ' ').replace('\r', '').replace('"', '').strip()
    #
    #             if cleaned:  # добавляем только если реально есть содержимое
    #                 col_values.append(cleaned)
    #
    #         # Склеиваем результат
    #         joined_string = "".join(col_values)
    #         ws["E5"].value = joined_string
    #
    #
    #         d5_value = user_data.get('d5', '').strip()
    #
    #         # Объединяем D5 (из интерфейса) и E5 (joined_string) в D5
    #         if d5_value and joined_string:
    #             ws["D5"].value = f"{d5_value} {joined_string}"
    #         elif d5_value:
    #             ws["D5"].value = d5_value
    #         elif joined_string:
    #             ws["D5"].value = joined_string
    #
    #         # Устанавливаем значения из user_data
    #         cell_mapping = {
    #             'b5': 'B5', 'c5': 'C5', 'f5': 'F5', 'h5': 'H5',
    #             'i5': 'I5', 'j5': 'J5', 'k5': 'K5', 'l5': 'L5',
    #             'm5': 'M5', 'n5': 'N5', 'o5': 'O5', 'p5': 'P5',
    #             'q5': 'Q5', 'r5': 'R5'
    #         }
    #
    #         for data_key, cell_ref in cell_mapping.items():
    #             value = user_data.get(data_key, '').strip()
    #             if value:
    #                 ws[cell_ref].value = value
    #
    #         sum_j = 0
    #         for row in range(3, 53):  # J3:J52
    #             raw_val = original_ws.cell(row=row, column=10).value  # колонка J = 10
    #             if raw_val is None:
    #                 continue
    #             try:
    #                 # Преобразуем в строку, убираем пробелы и "мм"
    #                 cleaned = str(raw_val).replace(" ", "").replace("мм", "").replace(",", ".").strip()
    #                 cleaned = re.sub(r"[^\d.]", "", cleaned)
    #                 if cleaned and cleaned != '.':  # ← Проверка на пустую строку и точку
    #                     number = float(cleaned)
    #
    #                     sum_j += number
    #                 if cleaned:
    #                     number = float(cleaned)
    #                     sum_j += number
    #             except Exception as err:
    #                 print(f"Ошибка в строке {row}: {err}")
    #                 continue
    #
    #         # Записываем сумму в G5
    #         ws["G5"].value = sum_j
    #
    #         # Закрашиваем G5 болотным цветом
    #         fill_olive = PatternFill(start_color="808000", end_color="808000", fill_type="solid")
    #         ws["G5"].fill = fill_olive
    #
    #         max_row = ws.max_row
    #         max_col = ws.max_column
    #
    #         for col in range(9, 12):
    #             ws.cell(row=4, column=col).value = "мм"
    #
    #         ws["A5"].value = 1
    #
    #         # O5
    #         try:
    #             g5_value = float(ws["G5"].value or 0)
    #             o5_value = float(user_data.get('o5', '0').replace(",", ".") or 0)
    #             ws["O5"].value = round(o5_value * g5_value / 1000, 3)
    #         except Exception as e:
    #             print("Ошибка при расчете O5:", e)
    #
    #         # R5
    #         try:
    #             g5_value = float(ws["G5"].value or 0)
    #             r5_value = float(user_data.get('r5', '0').replace(",", ".") or 0)
    #             ws["R5"].value = round(r5_value * g5_value / 1000, 3)
    #         except Exception as e:
    #             print("Ошибка при расчете R5:", e)
    #
    #         start_row = 3  # начинаем с 3-й строки исходного листа
    #         output_row = 3  # 3-я строка в целевом листе (там, где пишем значения)
    #         column_offset = 0  # сдвиг по столбцам, начиная со столбца S (19)
    #         empty_count = 0  # счётчик подряд пустых строк
    #
    #         while empty_count < 3:
    #             k_value = original_ws.cell(row=start_row, column=11).value  # L = 12
    #             l_value = original_ws.cell(row=start_row, column=12).value  # K = 13
    #
    #             if (k_value is None or str(l_value).strip() == "") and (l_value is None or str(k_value).strip() == ""):
    #                 empty_count += 1
    #             else:
    #                 l_str = str(l_value).strip() if l_value is not None else ""
    #                 k_str = str(k_value).strip() if k_value is not None else ""
    #                 combined = (k_str + " " + l_str).strip()
    #                 col_letter = get_column_letter(19 + column_offset)  # S = 19
    #                 ws[f"{col_letter}{output_row}"] = combined
    #
    #                 column_offset += 1
    #                 empty_count = 0  # сброс, т.к. строка не пустая
    #
    #             start_row += 1
    #
    #         empty_count = 0
    #         i = 3  # начинаем с F3
    #         output_row = 5
    #         column_offset = 0
    #
    #         while empty_count < 3:
    #             f_cell = original_ws.cell(row=i, column=5)  # F = 6
    #             f_value = f_cell.value
    #
    #             if f_value is None or str(f_value).strip() == "":
    #                 empty_count += 1
    #                 i += 1
    #                 continue
    #
    #             f_value_clean = str(f_value).strip().lower()
    #
    #             if f_value_clean == "кг":
    #                 value_to_write = original_ws.cell(row=i, column=9).value  # I = 9
    #                 empty_count = 0
    #             elif f_value_clean == "шт":
    #                 value_to_write = original_ws.cell(row=i, column=3).value  # C = 3
    #                 empty_count = 0
    #             else:
    #                 # если что-то другое (не "кг" и не "шт"), считаем как пустую
    #                 empty_count += 1
    #                 i += 1
    #                 continue
    #
    #             col_letter = get_column_letter(19 + column_offset)  # S = 19
    #             ws[f"{col_letter}{output_row}"] = value_to_write
    #
    #             column_offset += 1
    #             i += 1
    #
    #         # Заливка коричневым цветом
    #         fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
    #
    #         for offset in range(column_offset):
    #             col_letter = get_column_letter(19 + offset)  # 19 = S
    #
    #
    #         # Оформление + заголовки и типы
    #         for row_idx in range(1, max_row + 1):
    #             for col_idx in range(1, max_col + 1):
    #                 cell = ws.cell(row=row_idx, column=col_idx)
    #                 cell.border = border
    #
    #                 if row_idx == 1 and col_idx <= 19:
    #                     cell.fill = fill_blue  # A1:S1
    #
    #                 if row_idx in [2, 3, 4] and col_idx in [9, 10, 11]:
    #                     cell.fill = fill_blue  # J2:L3
    #
    #                 if row_idx == 1 and col_idx >= 19:
    #                     cell.fill = fill_pink  # S1 и далее
    #
    #                 if row_idx == 3 and col_idx >= 19:
    #                     cell.fill = fill_orange  # S3 и далее
    #
    #                 if row_idx == 4 and col_idx >= 19:
    #                     cell.fill = fill_darkblue  # S4 и далее
    #
    #                 if row_idx == 1 and col_idx >= 19:  # Изменяем на строку 1
    #                     top_cell_value = ws.cell(row=1, column=col_idx).value
    #                     if top_cell_value:
    #                         value_row1, value_row2 = self.process_value(top_cell_value)
    #                         # Записываем в строку 1 и строку 2
    #                         ws.cell(row=1, column=col_idx).value = value_row1
    #                         ws.cell(row=2, column=col_idx).value = value_row2
    #
    #         last_col = ws.max_column
    #
    #         # Цвет заливки — светло-синий
    #         light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    #
    #         # В первую ячейку последнего столбца пишем "Конец" и красим
    #         # ws.cell(row=1, column=last_col).value = "Конец"
    #
    #         # Обработка столбцов начиная с S (19)
    #         start_row = 3  # начинаем с 3-й строки исходного листа
    #         empty_count = 0
    #         column_offset = 0
    #
    #         for col_idx in range(19, max_col + 1):
    #             if col_idx <= ws.max_column:  # Проверяем что столбец существует
    #                 cell_4th_row = ws.cell(row=4, column=col_idx).value
    #                 cell_1st_row = ws.cell(row=1, column=col_idx).value
    #
    #                 # ЕСЛИ В 4-Й СТРОКЕ "ШТ" - АВТОМАТИЧЕСКОЕ ЗАПОЛНЕНИЕ
    #                 if cell_4th_row and str(cell_4th_row).strip().lower() == "шт" and cell_1st_row:
    #                     text_to_analyze = str(cell_1st_row).strip()
    #                     auto_value = ""
    #
    #                     # ТАБЛИЦА СООТВЕТСТВИЙ
    #                     if re.search(r'Фланец переходной.*\d+-\d+.*\d+-\d+', text_to_analyze):
    #                         auto_value = "09Г2С ФП-2024-КМД"
    #                     elif re.search(r'Фланец \d-\d+-\d+', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ 28759.2"
    #                     elif re.search(r'Фланец \d+-\d+-\d+-\d+-\w', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ33259-2015"
    #                     elif re.search(r'Крышка \d-\d+-\d+', text_to_analyze):
    #                         auto_value = "09Г2С ОСТ 26-2008-83"
    #                     elif re.search(r'Заглушка \d-\d+-\d+', text_to_analyze):
    #                         auto_value = "09Г2С АТК24.200.02-90"
    #                     elif re.search(r'Прокладка СНП', text_to_analyze):
    #                         auto_value = "ГОСТ Р 52376-2005"
    #                     elif re.search(r'Прокладка \d+-\w+', text_to_analyze):
    #                         auto_value = "ГОСТ 28759.6"
    #                     elif re.search(r'Прокладка \w-\d+-\d+-\w+', text_to_analyze):
    #                         auto_value = "ГОСТ 15180"
    #                     elif re.search(r'Прокладка +ПМБ-\d+', text_to_analyze):
    #                         auto_value = "ОСТ26.260.460-99"
    #                     elif re.search(r'(Бобышка БПО|Пробка)', text_to_analyze):
    #                         auto_value = "09Г2С ОСТ26.260.460-99"
    #                     elif re.search(r'Муфта \d+', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ 8966-75"
    #                     elif re.search(r'Сгон \d+', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ8969-75"
    #                     elif re.search(r'Ниппель \d+', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ8967-75"
    #                     elif re.search(r'Отвод', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ 17375-2001"
    #                     elif re.search(r'Переход', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ 17378-2001"
    #                     elif re.search(r'Тройник', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ 17376-2001"
    #                     elif re.search(r'Болт', text_to_analyze):
    #                         auto_value = "ГОСТ 7798-70"
    #                     elif re.search(r'Гайка', text_to_analyze):
    #                         auto_value = "ГОСТ 5915-70"
    #                     elif re.search(r'Шайба \d+', text_to_analyze):
    #                         auto_value = "ГОСТ 11371-78"
    #                     elif re.search(r'Шпилька', text_to_analyze):
    #                         auto_value = "09Г2С ГОСТ9066-75"
    #                     elif re.search(r'Шплинт', text_to_analyze):
    #                         auto_value = "ГОСТ 397-79"
    #                     elif re.search(r'Электроды', text_to_analyze):
    #                         auto_value = "ГОСТ9467-75"
    #                     elif re.search(r'(Скоба|Штырь)', text_to_analyze):
    #                         auto_value = "С245 ГОСТ17314-81"
    #
    #                     ws.cell(row=3, column=col_idx).value = auto_value
    #
    #                 # ИНАЧЕ - СТАРАЯ ЛОГИКА ПЕРЕНОСА ИЗ K И L
    #                 else:
    #                     k_value = original_ws.cell(row=start_row, column=11).value  # K = 11
    #                     l_value = original_ws.cell(row=start_row, column=12).value  # L = 12
    #
    #                     if (k_value is None or str(k_value).strip() == "") and (
    #                             l_value is None or str(l_value).strip() == ""):
    #                         empty_count += 1
    #                     else:
    #                         k_str = str(k_value).strip() if k_value is not None else ""
    #                         l_str = str(l_value).strip() if l_value is not None else ""
    #                         combined = (k_str + " " + l_str).strip()
    #
    #                         ws.cell(row=3, column=col_idx).value = combined
    #
    #                         column_offset += 1
    #                         empty_count = 0
    #
    #                     start_row += 1
    #
    #
    #
    #
    #         # 6. Копируем результат в output_ws
    #         for row in ws.iter_rows():
    #             for cell in row:
    #                 output_ws[cell.coordinate].value = cell.value
    #
    #         # 7. Очистка
    #         wb.close()
    #         original_wb.close()
    #         os.remove(temp_output)
    #
    #
    #
    #         max_row = output_ws.max_row
    #         max_col = output_ws.max_column
    #
    #         for row_idx in range(1, max_row + 1):
    #             for col_idx in range(1, max_col + 1):
    #                 cell = output_ws.cell(row=row_idx, column=col_idx)
    #                 cell.border = border
    #
    #
    #         return True
    #
    #     except Exception as e:
    #         print(f"Ошибка обработки листа {sheet_name}: {e}")
    #         return False


    def process_single_file(self, input_path, output_ws, user_data):
        """Обрабатывает один файл и сохраняет результат в указанный лист"""
        try:
            # 1. Загружаем и преобразуем данные (как в load_excel_from_path)
            df_raw = pd.read_excel(input_path, header=None)

            # 1. Загружаем и преобразуем данные (как в load_excel_from_path)
            df_raw = pd.read_excel(input_path, header=None)

            # ТВОЯ ЛОГИКА ПРЕОБРАЗОВАНИЯ DataFrame:
            new_header = [
                "№", "Тип Товара", "Наименование Товара", "Наименование для ВОМ",
                "Марка", "Нормативный документ", "Вес шт, кг", "Количество марок в единице Товара",
                "Длина", "Ширина", "Высота", "Вид ГМ", "Краска1", "Цвет1", "Расход краски1, кг",
                "Краска2", "Цвет2", "Расход краски2, кг"
            ]

            base_df = pd.DataFrame([new_header], columns=new_header)
            base_df = pd.concat([base_df, pd.DataFrame([[""] * len(new_header)] * 2, columns=new_header)],
                                ignore_index=True)

            l_vals = df_raw.iloc[2:, 11].dropna().tolist() if df_raw.shape[1] > 11 else []
            e_vals = df_raw.iloc[2:, 4].dropna().tolist() if df_raw.shape[1] > 4 else []
            f_vals = df_raw.iloc[2:, 5].dropna().tolist() if df_raw.shape[1] > 5 else []

            max_len = max(len(l_vals), len(e_vals), len(f_vals))

            while len(base_df.columns) < 19 + max_len:
                base_df[f"Доп.{len(base_df.columns) - 18}"] = ""

            current_df = base_df.copy()

            while len(current_df) < 4:
                current_df.loc[len(current_df)] = [""] * len(current_df.columns)

            for idx, val in enumerate(f_vals):
                current_df.iat[0, 18 + idx] = val  # строка 1

            for idx, val in enumerate(l_vals):
                current_df.iat[2, 18 + idx] = val  # строка 3

            for idx, val in enumerate(e_vals):
                current_df.iat[3, 18 + idx] = val  # строка 4

            # 2. Временно сохраняем преобразованный DataFrame
            temp_dir = tempfile.gettempdir()
            temp_output = os.path.join(temp_dir, f"temp_process_{os.getpid()}.xlsx")
            current_df.to_excel(temp_output, index=False, header=False)

            # 3. Открываем оригинальный файл для данных
            original_wb = load_workbook(input_path, data_only=True)
            original_ws = original_wb.active

            # 4. Открываем временный файл для оформления
            wb = load_workbook(temp_output)
            ws = wb.active

            # 5. ПРИМЕНЯЕМ ВСЮ ТВОЮ ЛОГИКУ ИЗ SAVE_EXCEL
            # (копируешь сюда весь код от "Цвета" до конца save_excel)

            # Цвета
            fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # светло-голубой
            fill_pink = PatternFill(start_color="D02090", end_color="D02090", fill_type="solid")  # малиновый
            fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # оранжевый
            fill_darkblue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")  # тёмно-голубой
            fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  # коричневый
            fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # фиолетовый
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # желтый

            # Чёрные границы
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            col_values = []
            for row in range(3, 23):  # от A3 до A22
                raw_value = original_ws.cell(row=row, column=1).value
                if raw_value is None:
                    continue

                # Преобразуем в строку и чистим от мусора
                cleaned = str(raw_value).replace('\n', ' ').replace('\r', '').replace('"', '').strip()

                if cleaned:  # добавляем только если реально есть содержимое
                    col_values.append(cleaned)

            # Склеиваем результат
            joined_string = "".join(col_values)
            wb = load_workbook(temp_output)
            ws = wb.active
            ws["E5"].value = joined_string


            d5_value = self.input_d5.currentText().strip()

            # Объединяем D5 (из интерфейса) и E5 (joined_string) в D5
            if d5_value and joined_string:
                ws["D5"].value = f"{d5_value} {joined_string}"  # Можно изменить разделитель
            elif d5_value:
                ws["D5"].value = d5_value
            elif joined_string:
                ws["D5"].value = joined_string

            b5_value = self.input_b5.currentText().strip()
            c5_value = self.input_c5.currentText().strip()
            f5_value = self.input_f5.currentText().strip()
            h5_value = self.input_h5.currentText().strip()
            i5_value = self.input_i5.currentText().strip()
            j5_value = self.input_j5.currentText().strip()
            k5_value = self.input_k5.currentText().strip()
            l5_value = self.input_l5.currentText().strip()
            m5_value = self.input_m5.currentText().strip()
            n5_value = self.input_n5.currentText().strip()
            o5_value = self.input_o5.currentText().strip()
            p5_value = self.input_p5.currentText().strip()
            q5_value = self.input_q5.currentText().strip()
            r5_value = self.input_r5.currentText().strip()

            if b5_value:
                ws["B5"].value = b5_value
            if c5_value:
                ws["C5"].value = c5_value
            if d5_value:
                ws["D5"].value = d5_value
            if f5_value:
                ws["F5"].value = f5_value
            if h5_value:
                ws["H5"].value = h5_value
            if i5_value:
                ws["I5"].value = i5_value
            if j5_value:
                ws["J5"].value = j5_value
            if k5_value:
                ws["K5"].value = k5_value
            if l5_value:
                ws["L5"].value = l5_value
            if m5_value:
                ws["M5"].value = m5_value
            if n5_value:
                ws["N5"].value = n5_value
            if o5_value:
                ws["O5"].value = o5_value
            if p5_value:
                ws["P5"].value = p5_value
            if q5_value:
                ws["Q5"].value = q5_value
            if r5_value:
                ws["R5"].value = r5_value

            sum_j = 0
            for row in range(3, 53):  # J3:J52
                raw_val = original_ws.cell(row=row, column=10).value  # колонка J = 10
                if raw_val is None:
                    continue
                try:
                    # Преобразуем в строку, убираем пробелы и "мм"
                    cleaned = str(raw_val).replace(" ", "").replace("мм", "").replace(",", ".").strip()
                    cleaned = re.sub(r"[^\d.]", "", cleaned)  # оставляем только цифры и точку

                    if cleaned:
                        number = float(cleaned)
                        sum_j += number
                except Exception as err:
                    print(f"Ошибка в строке {row}: {err}")
                    continue

            # Записываем сумму в G5
            ws["G5"].value = sum_j



            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(9, 12):
                ws.cell(row=4, column=col).value = "мм"

            ws["A5"].value = 1

            # O5
            try:
                g5_value = float(ws["G5"].value or 0)
                o5_value = float(o5_value.replace(",", ".") or 0)
                ws["O5"].value = round(o5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете O5:", e)

            # R5
            try:
                g5_value = float(ws["G5"].value or 0)
                r5_value = float(r5_value.replace(",", ".") or 0)
                ws["R5"].value = round(r5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете R5:", e)

            start_row = 3  # начинаем с 3-й строки исходного листа
            output_row = 3  # 3-я строка в целевом листе (там, где пишем значения)
            column_offset = 0  # сдвиг по столбцам, начиная со столбца S (19)
            empty_count = 0  # счётчик подряд пустых строк

            while empty_count < 3:
                k_value = original_ws.cell(row=start_row, column=11).value  # L = 12
                l_value = original_ws.cell(row=start_row, column=12).value  # K = 13

                if (k_value is None or str(l_value).strip() == "") and (l_value is None or str(k_value).strip() == ""):
                    empty_count += 1
                else:
                    l_str = str(l_value).strip() if l_value is not None else ""
                    k_str = str(k_value).strip() if k_value is not None else ""
                    combined = (k_str + " " + l_str).strip()
                    col_letter = get_column_letter(19 + column_offset)  # S = 19
                    ws[f"{col_letter}{output_row}"] = combined

                    column_offset += 1
                    empty_count = 0  # сброс, т.к. строка не пустая

                start_row += 1

            empty_count = 0
            i = 3  # начинаем с F3
            output_row = 5
            column_offset = 0

            while empty_count < 3:
                f_cell = original_ws.cell(row=i, column=5)  # F = 6
                f_value = f_cell.value

                if f_value is None or str(f_value).strip() == "":
                    empty_count += 1
                    i += 1
                    continue

                f_value_clean = str(f_value).strip().lower()

                if f_value_clean == "кг":
                    value_to_write = original_ws.cell(row=i, column=9).value  # I = 9
                    empty_count = 0
                elif f_value_clean == "шт":
                    value_to_write = original_ws.cell(row=i, column=3).value  # C = 3
                    empty_count = 0
                else:
                    # если что-то другое (не "кг" и не "шт"), считаем как пустую
                    empty_count += 1
                    i += 1
                    continue

                col_letter = get_column_letter(19 + column_offset)  # S = 19
                ws[f"{col_letter}{output_row}"] = value_to_write

                column_offset += 1
                i += 1

            # Заливка коричневым цветом
            fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")

            for offset in range(column_offset):
                col_letter = get_column_letter(19 + offset)  # 19 = S


            # Оформление + заголовки и типы
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border

                    if row_idx == 1 and col_idx >= 19:  # Изменяем на строку 1
                        top_cell_value = ws.cell(row=1, column=col_idx).value
                        if top_cell_value:
                            value_row1, value_row2 = self.process_value(top_cell_value)
                            # Записываем в строку 1 и строку 2
                            ws.cell(row=1, column=col_idx).value = value_row1
                            ws.cell(row=2, column=col_idx).value = value_row2


            last_col = ws.max_column

            # Цвет заливки — светло-синий
            light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # В первую ячейку последнего столбца пишем "Конец" и красим
            # ws.cell(row=1, column=last_col).value = "Конец"



            # 6. Копируем результат в output_ws - БЕЗ СТИЛЕЙ (сначала)
            print(f"Копируем данные из {os.path.basename(input_path)}")
            print(f"Размер данных: {ws.max_row} строк, {ws.max_column} столбцов")

            # Копируем только значения (без стилей)
            cell_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    output_ws[cell.coordinate].value = cell.value
                    if cell.value:
                        cell_count += 1

            print(f"Скопировано {cell_count} ячеек с данными")

            # 7. Очистка
            wb.close()
            original_wb.close()
            if os.path.exists(temp_output):
                os.remove(temp_output)
                print(f"Временный файл удален: {temp_output}")

            return True

        except Exception as e:
            print(f"Ошибка обработки {input_path}: {e}")
            return False

    def apply_full_styles(self, worksheet):
        """Применяет полную систему стилей как в save_excel"""
        try:
            # Все цвета из твоего save_excel
            fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            fill_pink = PatternFill(start_color="D02090", end_color="D02090", fill_type="solid")
            fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            fill_darkblue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
            fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            fill_olive = PatternFill(start_color="808000", end_color="808000", fill_type="solid")
            light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # Границы
            border = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )



            max_row = worksheet.max_row
            max_col = worksheet.max_column

            # 1. Применяем границы ко всем ячейкам
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = border

                    # 2. Обработка для строк 1 и 2 (столбцы S+)
            for col_idx in range(19, max_col + 1):
                if col_idx <= worksheet.max_column:
                    top_cell_value = worksheet.cell(row=1, column=col_idx).value
                if top_cell_value:
                    value_row1, value_row2 = self.process_value(top_cell_value)
                    worksheet.cell(row=1, column=col_idx).value = value_row1
                    worksheet.cell(row=2, column=col_idx).value = value_row2





            last_col = worksheet.max_column
            if last_col > 1:
                end_cell = worksheet.cell(row=1, column=last_col)
                if end_cell.value != "Конец":  # проверяем чтобы не дублировать
                    end_cell.value = "Конец"




            for col_idx in range(19, max_col + 1):
                if col_idx <= worksheet.max_column:
                    cell = worksheet.cell(row=5, column=col_idx)



            return True

        except Exception as e:
            print(f"Ошибка применения стилей: {e}")
            return False

    def universal_load(self):
        # Диалог выбора с опцией файлов и папок
        selected, _ = QFileDialog.getOpenFileNames(
            self,
            "Выберите файл(ы) или папку",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if not selected:
            return

        # Определяем тип выбора
        if len(selected) == 1 and os.path.isdir(selected[0]):
            self.process_folder(selected[0])
        elif len(selected) == 1:
            try:
                xl_file = pd.ExcelFile(selected[0])
                if len(xl_file.sheet_names) == 1:
                    # Загружаем и вызываем save_excel для диалога сохранения
                    if self.load_excel_file(selected[0], silent=True):
                        self.save_excel()  # ← ВЫЗЫВАЕМ СОХРАНЕНИЕ С ДИАЛОГОМ
                else:
                    self.process_multiple_files([selected[0]])
            except:
                if self.load_excel_file(selected[0], silent=True):
                    self.save_excel()  # ← ВЫЗЫВАЕМ СОХРАНЕНИЕ С ДИАЛОГОМ
        else:
            self.process_multiple_files(selected)

    def auto_save_excel(self):
        """Автоматически сохраняет файл после загрузки"""
        if self.df is None:
            return

        # Сохраняем рядом с исходным файлом
        if hasattr(self, 'last_loaded_path') and self.last_loaded_path:
            default_path = self.last_loaded_path.replace(".xlsx", "_обработанный.xlsx")
        else:
            default_path = os.path.join(os.path.expanduser("~"), "Desktop", "обработанный_файл.xlsx")

        # Можно сразу сохранить без диалога:
        try:
            self.save_excel_to_path(default_path)  # ← Нужно создать этот метод
            QMessageBox.information(self, "Сохранено", f"Файл сохранён:\n{default_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения: {str(e)}")

    def process_folder(self, folder_path):
        excel_files = []
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith(('.xlsx', '.xls')):
                    excel_files.append(os.path.join(root, file))

        if not excel_files:
            QMessageBox.information(self, "Информация", "В папке не найдено Excel-файлов")
            return

        self.process_multiple_files(excel_files)

    def clear_file_queue(self):
        """Очищает очередь файлов"""
        self.all_dropped_files = []
        self.update_file_status("Очередь очищена")
        self.label_queue_status.setText("Файлов в очереди: 0")

        # Обновляем DropFrame
        drop_frame = self.findChild(DropFrame)
        if drop_frame:
            drop_frame.setText("Перетащите Excel-файл сюда / папку с файлами")

        QMessageBox.information(self, "Очередь очищена", "Все файлы удалены из очереди")

    def process_all_files(self):
        """Обрабатывает все файлы в очереди"""
        if not self.all_dropped_files:
            QMessageBox.warning(self, "Очередь пуста", "Нет файлов для обработки")
            return

        if len(self.all_dropped_files) == 1:
            # Для одного файла проверяем количество листов
            file_path = self.all_dropped_files[0]
            try:
                xl_file = pd.ExcelFile(file_path)
                if len(xl_file.sheet_names) == 1:
                    # Загружаем и вызываем сохранение с диалогом
                    if self.load_excel_file(file_path, silent=True):
                        self.save_excel()  # ← ДОБАВЬ ДИАЛОГ СОХРАНЕНИЯ
                else:
                    # Многолистовой файл - обрабатываем как multiple
                    self.process_multiple_files(self.all_dropped_files)
            except:
                if self.load_excel_file(file_path, silent=True):
                    self.save_excel()  # ← ДОБАВЬ ДИАЛОГ СОХРАНЕНИЯ
        else:
            # Много файлов - обычная обработка
            self.process_multiple_files(self.all_dropped_files)

    def save_excel_to_path(self, file_path):

        try:
            # Шаг 1: Сохраняем DataFrame без заголовков
            self.df.to_excel(file_path, index=False, header=False)

            # Шаг 2: Открываем сохранённый файл
            # Открываем именно исходный файл, а не новый!
            original_wb = load_workbook(self.last_loaded_path, data_only=True)
            original_ws = original_wb.active


            # Цвета
            fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # светло-голубой
            fill_pink = PatternFill(start_color="D02090", end_color="D02090", fill_type="solid")  # малиновый
            fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # оранжевый
            fill_darkblue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")  # тёмно-голубой
            fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  # коричневый
            fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # фиолетовый
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # желтый

            # Чёрные границы
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )



            col_values = []
            for row in range(3, 23):  # от A3 до A22
                raw_value = original_ws.cell(row=row, column=1).value
                if raw_value is None:
                    continue

                # Преобразуем в строку и чистим от мусора
                cleaned = str(raw_value).replace('\n', ' ').replace('\r', '').replace('"', '').strip()

                if cleaned:  # добавляем только если реально есть содержимое
                    col_values.append(cleaned)



            # Склеиваем результат
            joined_string = "".join(col_values)
            wb = load_workbook(file_path)
            ws = wb.active
            ws["E5"].value = joined_string


            d5_value = self.input_d5.currentText().strip()

            # Объединяем D5 (из интерфейса) и E5 (joined_string) в D5
            if d5_value and joined_string:
                ws["D5"].value = f"{d5_value} {joined_string}"  # Можно изменить разделитель
            elif d5_value:
                ws["D5"].value = d5_value
            elif joined_string:
                ws["D5"].value = joined_string

            b5_value = self.input_b5.currentText().strip()
            c5_value = self.input_c5.currentText().strip()
            f5_value = self.input_f5.currentText().strip()
            h5_value = self.input_h5.currentText().strip()
            i5_value = self.input_i5.currentText().strip()
            j5_value = self.input_j5.currentText().strip()
            k5_value = self.input_k5.currentText().strip()
            l5_value = self.input_l5.currentText().strip()
            m5_value = self.input_m5.currentText().strip()
            n5_value = self.input_n5.currentText().strip()
            o5_value = self.input_o5.currentText().strip()
            p5_value = self.input_p5.currentText().strip()
            q5_value = self.input_q5.currentText().strip()
            r5_value = self.input_r5.currentText().strip()



            if b5_value:
                ws["B5"].value = b5_value
            if c5_value:
                ws["C5"].value = c5_value
            if d5_value:
                ws["D5"].value = d5_value
            if f5_value:
                ws["F5"].value = f5_value
            if h5_value:
                ws["H5"].value = h5_value
            if i5_value:
                ws["I5"].value = i5_value
            if j5_value:
                ws["J5"].value = j5_value
            if k5_value:
                ws["K5"].value = k5_value
            if l5_value:
                ws["L5"].value = l5_value
            if m5_value:
                ws["M5"].value = m5_value
            if n5_value:
                ws["N5"].value = n5_value
            if o5_value:
                ws["O5"].value = o5_value
            if p5_value:
                ws["P5"].value = p5_value
            if q5_value:
                ws["Q5"].value = q5_value
            if r5_value:
                ws["R5"].value = r5_value

            sum_j = 0
            for row in range(3, 53):  # J3:J52
                raw_val = original_ws.cell(row=row, column=10).value  # колонка J = 10
                if raw_val is None:
                    continue
                try:
                    # Преобразуем в строку, убираем пробелы и "мм"
                    cleaned = str(raw_val).replace(" ", "").replace("мм", "").replace(",", ".").strip()
                    cleaned = re.sub(r"[^\d.]", "", cleaned)  # оставляем только цифры и точку

                    if cleaned:
                        number = float(cleaned)
                        sum_j += number
                except Exception as err:
                    print(f"Ошибка в строке {row}: {err}")
                    continue

            # Записываем сумму в G5
            ws["G5"].value = sum_j



            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(9, 12):
                ws.cell(row=4, column=col).value = "мм"

            ws["A5"].value = 1

            # O5
            try:
                g5_value = float(ws["G5"].value or 0)
                o5_value = float(o5_value.replace(",", ".") or 0)
                ws["O5"].value = round(o5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете O5:", e)

            # R5
            try:
                g5_value = float(ws["G5"].value or 0)
                r5_value = float(r5_value.replace(",", ".") or 0)
                ws["R5"].value = round(r5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете R5:", e)

            start_row = 3  # начинаем с 3-й строки исходного листа
            output_row = 3  # 3-я строка в целевом листе (там, где пишем значения)
            column_offset = 0  # сдвиг по столбцам, начиная со столбца S (19)
            empty_count = 0  # счётчик подряд пустых строк

            while empty_count < 3:
                k_value = original_ws.cell(row=start_row, column=11).value  # L = 12
                l_value = original_ws.cell(row=start_row, column=12).value  # K = 13

                if (k_value is None or str(l_value).strip() == "") and (l_value is None or str(k_value).strip() == ""):
                    empty_count += 1
                else:
                    l_str = str(l_value).strip() if l_value is not None else ""
                    k_str = str(k_value).strip() if k_value is not None else ""
                    combined = (k_str + " " + l_str).strip()
                    col_letter = get_column_letter(19 + column_offset)  # S = 19
                    ws[f"{col_letter}{output_row}"] = combined

                    column_offset += 1
                    empty_count = 0  # сброс, т.к. строка не пустая

                start_row += 1

            empty_count = 0
            i = 3  # начинаем с F3
            output_row = 5
            column_offset = 0

            while empty_count < 3:
                f_cell = original_ws.cell(row=i, column=5)  # F = 6
                f_value = f_cell.value

                if f_value is None or str(f_value).strip() == "":
                    empty_count += 1
                    i += 1
                    continue

                f_value_clean = str(f_value).strip().lower()

                if f_value_clean == "кг":
                    value_to_write = original_ws.cell(row=i, column=9).value  # I = 9
                    empty_count = 0
                elif f_value_clean == "шт":
                    value_to_write = original_ws.cell(row=i, column=3).value  # C = 3
                    empty_count = 0
                else:
                    # если что-то другое (не "кг" и не "шт"), считаем как пустую
                    empty_count += 1
                    i += 1
                    continue

                col_letter = get_column_letter(19 + column_offset)  # S = 19
                ws[f"{col_letter}{output_row}"] = value_to_write

                column_offset += 1
                i += 1



            for offset in range(column_offset):
                col_letter = get_column_letter(19 + offset)  # 19 = S






            # Оформление + заголовки и типы
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border

                    if row_idx == 1 and col_idx >= 19:  # Изменяем на строку 1
                        top_cell_value = ws.cell(row=1, column=col_idx).value
                        if top_cell_value:
                            value_row1, value_row2 = self.process_value(top_cell_value)
                            # Записываем в строку 1 и строку 2
                            ws.cell(row=1, column=col_idx).value = value_row1
                            ws.cell(row=2, column=col_idx).value = value_row2


            last_col = ws.max_column

            # Цвет заливки — светло-синий
            light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # В первую ячейку последнего столбца пишем "Конец" и красим
            ws.cell(row=1, column=last_col).value = "Конец"






            wb.save(file_path)
            QMessageBox.information(self, "Сохранено", f"Файл сохранён с оформлением:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка при сохранении", str(e))




    def save_excel(self):
        if self.df is None:
            QMessageBox.warning(self, "Нет данных", "Сначала загрузите Excel-файл.")
            return

            # Если last_loaded_path есть, предлагаем сохранить рядом с ним с суффиксом
        if hasattr(self, 'last_loaded_path') and self.last_loaded_path:
            default_path = self.last_loaded_path.replace(".xlsx", "_обработанный.xlsx")
        else:
            default_path = os.path.join(os.path.expanduser("~"), "Desktop", "обработанный_файл.xlsx")

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel-файл", default_path, "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            # Шаг 1: Сохраняем DataFrame без заголовков
            self.df.to_excel(file_path, index=False, header=False)

            # Шаг 2: Открываем сохранённый файл
            # Открываем именно исходный файл, а не новый!
            original_wb = load_workbook(self.last_loaded_path, data_only=True)
            original_ws = original_wb.active


            # Цвета
            fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # светло-голубой
            fill_pink = PatternFill(start_color="D02090", end_color="D02090", fill_type="solid")  # малиновый
            fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # оранжевый
            fill_darkblue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")  # тёмно-голубой
            fill_brown = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  # коричневый
            fill_purple = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # фиолетовый
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # желтый

            # Чёрные границы
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )



            col_values = []
            for row in range(3, 23):  # от A3 до A22
                raw_value = original_ws.cell(row=row, column=1).value
                if raw_value is None:
                    continue

                # Преобразуем в строку и чистим от мусора
                cleaned = str(raw_value).replace('\n', ' ').replace('\r', '').replace('"', '').strip()

                if cleaned:  # добавляем только если реально есть содержимое
                    col_values.append(cleaned)



            # Склеиваем результат
            joined_string = "".join(col_values)
            wb = load_workbook(file_path)
            ws = wb.active
            ws["E5"].value = joined_string


            d5_value = self.input_d5.currentText().strip()

            # Объединяем D5 (из интерфейса) и E5 (joined_string) в D5
            if d5_value and joined_string:
                ws["D5"].value = f"{d5_value} {joined_string}"  # Можно изменить разделитель
            elif d5_value:
                ws["D5"].value = d5_value
            elif joined_string:
                ws["D5"].value = joined_string

            b5_value = self.input_b5.currentText().strip()
            c5_value = self.input_c5.currentText().strip()
            f5_value = self.input_f5.currentText().strip()
            h5_value = self.input_h5.currentText().strip()
            i5_value = self.input_i5.currentText().strip()
            j5_value = self.input_j5.currentText().strip()
            k5_value = self.input_k5.currentText().strip()
            l5_value = self.input_l5.currentText().strip()
            m5_value = self.input_m5.currentText().strip()
            n5_value = self.input_n5.currentText().strip()
            o5_value = self.input_o5.currentText().strip()
            p5_value = self.input_p5.currentText().strip()
            q5_value = self.input_q5.currentText().strip()
            r5_value = self.input_r5.currentText().strip()



            if b5_value:
                ws["B5"].value = b5_value
            if c5_value:
                ws["C5"].value = c5_value
            if d5_value:
                ws["D5"].value = d5_value
            if f5_value:
                ws["F5"].value = f5_value
            if h5_value:
                ws["H5"].value = h5_value
            if i5_value:
                ws["I5"].value = i5_value
            if j5_value:
                ws["J5"].value = j5_value
            if k5_value:
                ws["K5"].value = k5_value
            if l5_value:
                ws["L5"].value = l5_value
            if m5_value:
                ws["M5"].value = m5_value
            if n5_value:
                ws["N5"].value = n5_value
            if o5_value:
                ws["O5"].value = o5_value
            if p5_value:
                ws["P5"].value = p5_value
            if q5_value:
                ws["Q5"].value = q5_value
            if r5_value:
                ws["R5"].value = r5_value

            sum_j = 0
            for row in range(3, 53):  # J3:J52
                raw_val = original_ws.cell(row=row, column=10).value  # колонка J = 10
                if raw_val is None:
                    continue
                try:
                    # Преобразуем в строку, убираем пробелы и "мм"
                    cleaned = str(raw_val).replace(" ", "").replace("мм", "").replace(",", ".").strip()
                    cleaned = re.sub(r"[^\d.]", "", cleaned)  # оставляем только цифры и точку

                    if cleaned:
                        number = float(cleaned)
                        sum_j += number
                except Exception as err:
                    print(f"Ошибка в строке {row}: {err}")
                    continue

            # Записываем сумму в G5
            ws["G5"].value = sum_j



            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(9, 12):
                ws.cell(row=4, column=col).value = "мм"

            ws["A5"].value = 1

            # O5
            try:
                g5_value = float(ws["G5"].value or 0)
                o5_value = float(o5_value.replace(",", ".") or 0)
                ws["O5"].value = round(o5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете O5:", e)

            # R5
            try:
                g5_value = float(ws["G5"].value or 0)
                r5_value = float(r5_value.replace(",", ".") or 0)
                ws["R5"].value = round(r5_value * g5_value / 1000, 3)
            except Exception as e:
                print("Ошибка при расчете R5:", e)

            start_row = 3  # начинаем с 3-й строки исходного листа
            output_row = 3  # 3-я строка в целевом листе (там, где пишем значения)
            column_offset = 0  # сдвиг по столбцам, начиная со столбца S (19)
            empty_count = 0  # счётчик подряд пустых строк

            while empty_count < 3:
                k_value = original_ws.cell(row=start_row, column=11).value  # L = 12
                l_value = original_ws.cell(row=start_row, column=12).value  # K = 13

                if (k_value is None or str(l_value).strip() == "") and (l_value is None or str(k_value).strip() == ""):
                    empty_count += 1
                else:
                    l_str = str(l_value).strip() if l_value is not None else ""
                    k_str = str(k_value).strip() if k_value is not None else ""
                    combined = (k_str + " " + l_str).strip()
                    col_letter = get_column_letter(19 + column_offset)  # S = 19
                    ws[f"{col_letter}{output_row}"] = combined

                    column_offset += 1
                    empty_count = 0  # сброс, т.к. строка не пустая

                start_row += 1

            empty_count = 0
            i = 3  # начинаем с F3
            output_row = 5
            column_offset = 0

            while empty_count < 3:
                f_cell = original_ws.cell(row=i, column=5)  # F = 6
                f_value = f_cell.value

                if f_value is None or str(f_value).strip() == "":
                    empty_count += 1
                    i += 1
                    continue

                f_value_clean = str(f_value).strip().lower()

                if f_value_clean == "кг":
                    value_to_write = original_ws.cell(row=i, column=9).value  # I = 9
                    empty_count = 0
                elif f_value_clean == "шт":
                    value_to_write = original_ws.cell(row=i, column=3).value  # C = 3
                    empty_count = 0
                else:
                    # если что-то другое (не "кг" и не "шт"), считаем как пустую
                    empty_count += 1
                    i += 1
                    continue

                col_letter = get_column_letter(19 + column_offset)  # S = 19
                ws[f"{col_letter}{output_row}"] = value_to_write

                column_offset += 1
                i += 1

            # Обработка столбцов начиная с S (19)
            start_row = 3  # начинаем с 3-й строки исходного листа
            empty_count = 0
            column_offset = 0

            for col_idx in range(19, max_col + 1):
                if col_idx <= ws.max_column:  # Проверяем что столбец существует
                    cell_4th_row = ws.cell(row=4, column=col_idx).value
                    cell_1st_row = ws.cell(row=1, column=col_idx).value

                    # ЕСЛИ В 4-Й СТРОКЕ "ШТ" - АВТОМАТИЧЕСКОЕ ЗАПОЛНЕНИЕ
                    if cell_4th_row and str(cell_4th_row).strip().lower() == "шт" and cell_1st_row:
                        text_to_analyze = str(cell_1st_row).strip()
                        auto_value = ""

                        # ТАБЛИЦА СООТВЕТСТВИЙ
                        if re.search(r'Фланец переходной.*\d+-\d+.*\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С ФП-2024-КМД"
                        elif re.search(r'Фланец \d-\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 28759.2"
                        elif re.search(r'Фланец \d+-\d+-\d+-\d+-\w', text_to_analyze):
                            auto_value = "09Г2С ГОСТ33259-2015"
                        elif re.search(r'Крышка \d-\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С ОСТ 26-2008-83"
                        elif re.search(r'Заглушка \d-\d+-\d+', text_to_analyze):
                            auto_value = "09Г2С АТК24.200.02-90"
                        elif re.search(r'Прокладка СНП', text_to_analyze):
                            auto_value = "ГОСТ Р 52376-2005"
                        elif re.search(r'Прокладка \d+-\w+', text_to_analyze):
                            auto_value = "ГОСТ 28759.6"
                        elif re.search(r'Прокладка \w-\d+-\d+-\w+', text_to_analyze):
                            auto_value = "ГОСТ 15180"
                        elif re.search(r'Прокладка \d+-ПМБ', text_to_analyze):
                            auto_value = "ОСТ26.260.460-99"
                        elif re.search(r'(Бобышка БПО|Пробка)', text_to_analyze):
                            auto_value = "09Г2С ОСТ26.260.460-99"
                        elif re.search(r'Муфта \d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 8966-75"
                        elif re.search(r'Сгон \d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ8969-75"
                        elif re.search(r'Ниппель \d+', text_to_analyze):
                            auto_value = "09Г2С ГОСТ8967-75"
                        elif re.search(r'Отвод', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 17375-2001"
                        elif re.search(r'Переход', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 17378-2001"
                        elif re.search(r'Тройник', text_to_analyze):
                            auto_value = "09Г2С ГОСТ 17376-2001"
                        elif re.search(r'Болт', text_to_analyze):
                            auto_value = "ГОСТ 7798-70"
                        elif re.search(r'Гайка', text_to_analyze):
                            auto_value = "ГОСТ 5915-70"
                        elif re.search(r'Шайба \d+', text_to_analyze):
                            auto_value = "ГОСТ 11371-78"
                        elif re.search(r'Шпилька', text_to_analyze):
                            auto_value = "09Г2С ГОСТ9066-75"
                        elif re.search(r'Шплинт', text_to_analyze):
                            auto_value = "ГОСТ 397-79"
                        elif re.search(r'Электроды', text_to_analyze):
                            auto_value = "ГОСТ9467-75"
                        elif re.search(r'(Скоба|Штырь)', text_to_analyze):
                            auto_value = "С245 ГОСТ17314-81"

                        ws.cell(row=3, column=col_idx).value = auto_value

                    # ИНАЧЕ - СТАРАЯ ЛОГИКА ПЕРЕНОСА ИЗ K И L
                    else:
                        k_value = original_ws.cell(row=start_row, column=11).value  # K = 11
                        l_value = original_ws.cell(row=start_row, column=12).value  # L = 12

                        if (k_value is None or str(k_value).strip() == "") and (
                                l_value is None or str(l_value).strip() == ""):
                            empty_count += 1
                        else:
                            k_str = str(k_value).strip() if k_value is not None else ""
                            l_str = str(l_value).strip() if l_value is not None else ""
                            combined = (k_str + " " + l_str).strip()

                            ws.cell(row=3, column=col_idx).value = combined

                            column_offset += 1
                            empty_count = 0

                        start_row += 1


            for offset in range(column_offset):
                col_letter = get_column_letter(19 + offset)  # 19 = S






            # Оформление + заголовки и типы
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border

                    if row_idx == 1 and col_idx >= 19:  # Изменяем на строку 1
                        top_cell_value = ws.cell(row=1, column=col_idx).value
                        if top_cell_value:
                            value_row1, value_row2 = self.process_value(top_cell_value)
                            # Записываем в строку 1 и строку 2
                            ws.cell(row=1, column=col_idx).value = value_row1
                            ws.cell(row=2, column=col_idx).value = value_row2

            last_col = ws.max_column

            # Цвет заливки — светло-синий
            light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # В первую ячейку последнего столбца пишем "Конец" и красим
            ws.cell(row=1, column=last_col).value = "Конец"






            wb.save(file_path)
            QMessageBox.information(self, "Сохранено", f"Файл сохранён с оформлением:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка при сохранении", str(e))








class FileProcessingThread(QThread):
    progress = pyqtSignal(int, str)  # progress_index, filename
    finished = pyqtSignal(bool, str)  # success, message
    error = pyqtSignal(str)  # error_message

    def __init__(self, file_paths, user_data, parent=None):
        super().__init__(parent)
        self.file_paths = file_paths
        self.user_data = user_data
        self.output_path = ""


    def run(self):
        try:
            from openpyxl import Workbook
            result_wb = Workbook()
            result_wb.remove(result_wb.active)
            processed_count = 0

            for i, file_path in enumerate(self.file_paths):
                if self.isInterruptionRequested():
                    break

                self.progress.emit(i + 1, os.path.basename(file_path))

                try:
                    # Обрабатываем все листы в файле
                    xl_file = pd.ExcelFile(file_path)
                    file_name = os.path.splitext(os.path.basename(file_path))[0]

                    for sheet_name in xl_file.sheet_names:
                        safe_sheet_name = f"{file_name}_{sheet_name}"[:31]
                        result_ws = result_wb.create_sheet(title=safe_sheet_name)

                        # Обрабатываем лист (пока заглушка)
                        success = self.process_sheet(file_path, sheet_name, result_ws)
                        if success:
                            processed_count += 1

                except Exception as e:
                    self.error.emit(f"Ошибка в {os.path.basename(file_path)}: {str(e)}")

            # Сохраняем результат
            output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            self.output_path = os.path.join(output_dir, "обработанные_файлы.xlsx")
            result_wb.save(self.output_path)

            self.finished.emit(True, f"Обработано {processed_count} листов\nСохранено в: {self.output_path}")

        except Exception as e:
            self.finished.emit(False, f"Критическая ошибка: {str(e)}")

    def process_sheet(self, file_path, sheet_name, target_ws):
        """Заглушка для обработки листа - позже заменим на вашу логику"""
        try:
            # Пока просто копируем данные
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    target_ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

            return True
        except Exception as e:
            print(f"Ошибка обработки листа {sheet_name}: {e}")
            return False












if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelApp()
    window.show()
    sys.exit(app.exec())
